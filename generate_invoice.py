# generate_invoice.py
# Main script to orchestrate invoice generation using templates, config files, and utility functions.
# Updated: Handles multi-table processing and configurable header writing/finding.
# Updated: Added support for loading data from .pkl (Pickle) files.
# Updated: Removed fallback to find_header; requires write_header config for single-table sheets.
# Updated: Adds a merged blank row between multi-table sections (e.g., Packing List).
# Updated: Leaves a single blank row after the last table based on config 'row_spacing' (removed final merge).
# REVISED AGAIN: Corrected flag retrieval and passing in generate_invoice.py. Includes final debug prints.
# Updated: Added configurable merging for rows adjacent to header/footer.
# Updated: Integrated configurable styling (font, alignment, widths) from config file.
# Updated: Integrated pallet order tracking across multi-table chunks.
# MODIFIED: Calculates final_grand_total_pallets globally before sheet loop and passes it to all fill_invoice_data calls.

import os
import json
import pickle # Import pickle module
import argparse
import shutil
import openpyxl
import traceback
import sys
from pathlib import Path
from typing import Optional, Dict, Any, Union, List, Tuple
import ast # <-- Add import for literal_eval
from decimal import Decimal # <-- Add import for Decimal evaluation
import re # <-- Add import for regular expressions
from openpyxl.utils import get_column_letter # REMOVED range_boundaries
import text_replace_utils # Ensure this is imported

# --- Import utility functions ---
try:
    # Ensure invoice_utils.py corresponds to the latest version with pallet order updates
    import invoice_utils
    import merge_utils # <-- Import the new merge utility module
    print("Successfully imported invoice_utils and merge_utils.")
except ImportError as import_err:
    print("------------------------------------------------------")
    print(f"FATAL ERROR: Could not import required utility modules: {import_err}")
    print("Please ensure invoice_utils.py and merge_utils.py are in the same directory as generate_invoice.py.")
    print("------------------------------------------------------")
    sys.exit(1)

# --- Helper Functions (derive_paths, load_config, load_data) ---
# Assume these functions exist as previously defined in the uploaded file.
# They are omitted here for brevity but are required for the script to work.
# Make sure to include the full code for these functions from your original file.
# --- Placeholder for Required Helper Functions ---
# NOTE: Replace 'pass' with the actual function definitions from your original file
def derive_paths(input_data_path_str: str, template_dir_str: str, config_dir_str: str) -> Optional[Dict[str, Path]]:
    """
    Derives template and config file paths based on the input data filename.
    Checks if template/config paths are valid directories.
    Assumes data file is named like TEMPLATE_NAME.xxx or TEMPLATE_NAME_data.xxx
    Attempts prefix matching if exact match is not found.
    """
    print(f"Deriving paths from input: {input_data_path_str}")
    try:
        input_data_path = Path(input_data_path_str).resolve()
        template_dir = Path(template_dir_str).resolve()
        config_dir = Path(config_dir_str).resolve()

        if not input_data_path.is_file(): print(f"Error: Input data file not found: {input_data_path}"); return None
        if not template_dir.is_dir(): print(f"Error: Template directory not found: {template_dir}"); return None
        if not config_dir.is_dir(): print(f"Error: Config directory not found: {config_dir}"); return None

        base_name = input_data_path.stem
        template_name_part = base_name
        suffixes_to_remove = ['_data', '_input', '_pkl']
        prefixes_to_remove = ['data_']

        for suffix in suffixes_to_remove:
            if base_name.lower().endswith(suffix):
                template_name_part = base_name[:-len(suffix)]
                break
        else: # Only check prefixes if no suffix was removed
            for prefix in prefixes_to_remove:
                if base_name.lower().startswith(prefix):
                    template_name_part = base_name[len(prefix):]
                    break

        if not template_name_part:
            print(f"Error: Could not derive template name part from: '{base_name}'")
            return None
        print(f"Derived initial template name part: '{template_name_part}'")

        # --- Attempt 1: Exact Match ---
        exact_template_filename = f"{template_name_part}.xlsx"
        exact_config_filename = f"{template_name_part}_config.json"
        exact_template_path = template_dir / exact_template_filename
        exact_config_path = config_dir / exact_config_filename
        print(f"Checking for exact match: Template='{exact_template_path}', Config='{exact_config_path}'")

        if exact_template_path.is_file() and exact_config_path.is_file():
            print("Found exact match for template and config.")
            return {"data": input_data_path, "template": exact_template_path, "config": exact_config_path}
        else:
            print("Exact match not found. Attempting prefix matching...")

            # --- Attempt 2: Prefix Match ---
            prefix_match = re.match(r'^([a-zA-Z]+)', template_name_part) # Extract leading letters
            if prefix_match:
                prefix = prefix_match.group(1)
                print(f"Extracted prefix: '{prefix}'")
                prefix_template_filename = f"{prefix}.xlsx"
                prefix_config_filename = f"{prefix}_config.json"
                prefix_template_path = template_dir / prefix_template_filename
                prefix_config_path = config_dir / prefix_config_filename
                print(f"Checking for prefix match: Template='{prefix_template_path}', Config='{prefix_config_path}'")

                if prefix_template_path.is_file() and prefix_config_path.is_file():
                    print("Found prefix match for template and config.")
                    return {"data": input_data_path, "template": prefix_template_path, "config": prefix_config_path}
                else:
                    print("Prefix match not found.")
            else:
                print("Could not extract a letter-based prefix.")

            # --- No Match Found ---
            print(f"Error: Could not find matching template/config files using exact ('{template_name_part}') or prefix methods.")
            # Report specific missing files based on the exact match attempt
            if not exact_template_path.is_file(): print(f"Error: Template file not found: {exact_template_path}")
            if not exact_config_path.is_file(): print(f"Error: Configuration file not found: {exact_config_path}")
            return None

    except Exception as e:
        print(f"Error deriving file paths: {e}")
        traceback.print_exc()
        return None

def load_config(config_path: Path) -> Optional[Dict[str, Any]]:
    """Loads and parses the JSON configuration file."""
    print(f"Loading configuration from: {config_path}")
    try:
        with open(config_path, 'r', encoding='utf-8') as f: config_data = json.load(f)
        print("Configuration loaded successfully.")
        if not isinstance(config_data, dict): print("Error: Config file is not a valid JSON object."); return None
        # Basic validation (add checks for 'styling' if required globally, but usually per-sheet)
        required_keys = ['sheets_to_process', 'sheet_data_map', 'data_mapping']
        missing_keys = [key for key in required_keys if key not in config_data]
        if missing_keys: print(f"Error: Config file missing required keys: {', '.join(missing_keys)}"); return None
        if not isinstance(config_data.get('data_mapping'), dict): print(f"Error: 'data_mapping' section is not a valid dictionary."); return None
        return config_data
    except json.JSONDecodeError as e: print(f"Error: Invalid JSON in configuration file {config_path}: {e}"); return None
    except Exception as e: print(f"Error loading configuration file {config_path}: {e}"); traceback.print_exc(); return None

def load_data(data_path: Path) -> Optional[Dict[str, Any]]:
    """ Loads and parses the input data file. Supports .json and .pkl. """
    print(f"Loading data from: {data_path}")
    invoice_data = None; file_suffix = data_path.suffix.lower()
    try:
        if file_suffix == '.json':
            print("Detected .json file...")
            with open(data_path, 'r', encoding='utf-8') as f: invoice_data = json.load(f)
            print("JSON data loaded successfully.")
        elif file_suffix == '.pkl':
            print("Detected .pkl file...");
            with open(data_path, 'rb') as f: invoice_data = pickle.load(f)
            print("Pickle data loaded successfully.")
        else: print(f"Error: Unsupported data file extension: '{file_suffix}'."); return None
        if not isinstance(invoice_data, dict): print("Error: Loaded data is not a dictionary."); return None

        # --- START AGGREGATION KEY CONVERSION ---
        # Use "initial_standard_aggregation" as requested
        aggregation_data_raw = invoice_data.get("standard_aggregation_results")
        if isinstance(aggregation_data_raw, dict):
            print("DEBUG: Found 'standard_aggregation_results'. Converting string keys to tuples...")
            aggregation_data_processed = {}
            converted_count = 0
            conversion_errors = 0
            # Regex to find Decimal('...') and capture the inner number string
            decimal_pattern = re.compile(r"Decimal\('(-?\d*\.?\d+)'\)") # Handles optional -, digits, optional decimal point

            for key_str, value_dict in aggregation_data_raw.items():
                processed_key_str = key_str # Initialize for error message
                try:
                    # Preprocess the string: Replace Decimal('...') with just the number string '...'
                    processed_key_str = decimal_pattern.sub(r"'\1'", key_str) # Replace with the number in quotes

                    # Now evaluate the processed string which should only contain literals
                    key_tuple = ast.literal_eval(processed_key_str)

                    # --- START MODIFIED POST-PROCESSING ---
                    # Convert tuple elements: Keep PO (idx 0) and Item (idx 1) as strings,
                    # convert Unit Price (idx 2) to float.
                    final_key_list = []
                    if isinstance(key_tuple, tuple) and len(key_tuple) >= 3:
                        # PO Number (Index 0): Keep as string
                        final_key_list.append(str(key_tuple[0]))

                        # Item Number (Index 1): Keep as string
                        final_key_list.append(str(key_tuple[1]))

                        # Unit Price (Index 2): Convert to float
                        unit_price_val = key_tuple[2]
                        if isinstance(unit_price_val, (int, float)):
                            final_key_list.append(float(unit_price_val))
                        elif isinstance(unit_price_val, str):
                            try:
                                final_key_list.append(float(unit_price_val))
                            except ValueError:
                                print(f"Warning: Could not convert unit price string '{unit_price_val}' to float for key '{key_str}'. Keeping as string.")
                                final_key_list.append(unit_price_val) # Keep original string on error
                        else:
                            # Handle other types if necessary, maybe try converting to float
                            try: final_key_list.append(float(unit_price_val))
                            except (ValueError, TypeError):
                                print(f"Warning: Could not convert unit price type '{type(unit_price_val)}' ({unit_price_val}) to float for key '{key_str}'. Keeping original type.")
                                final_key_list.append(unit_price_val) # Keep original on error

                        # Add any remaining elements from the original tuple (if any)
                        if len(key_tuple) > 3:
                            final_key_list.extend(key_tuple[3:])

                    else:
                        # Handle cases where the tuple doesn't have the expected structure
                        print(f"Warning: Evaluated key tuple '{key_tuple}' does not have expected length >= 3. Using original items.")
                        final_key_list = list(key_tuple) # Use original items

                    final_key_tuple = tuple(final_key_list)
                    # --- END MODIFIED POST-PROCESSING ---


                    if isinstance(final_key_tuple, tuple):
                        aggregation_data_processed[final_key_tuple] = value_dict
                        converted_count += 1
                    else:
                        # This case should be less likely now with the explicit tuple check above
                        print(f"Warning: Final key is not a tuple for processed key string '{processed_key_str}'. Original: '{key_str}'. Result: {final_key_tuple}")
                        conversion_errors += 1
                except (ValueError, SyntaxError, NameError, TypeError) as e:
                    print(f"Warning: Could not convert aggregation key string '{key_str}' (processed: '{processed_key_str}') to tuple: {e}")
                    conversion_errors += 1
            # Replace the original string-keyed dict with the tuple-keyed one
            # Update the key used for replacement as well
            invoice_data["standard_aggregation_results"] = aggregation_data_processed
            print(f"DEBUG: Finished key conversion. Converted: {converted_count}, Errors: {conversion_errors}")
        # --- END AGGREGATION KEY CONVERSION ---

        # --- START CUSTOM AGGREGATION KEY CONVERSION ---
        # Added block to handle custom_aggregation_results
        custom_aggregation_data_raw = invoice_data.get("custom_aggregation_results")
        if isinstance(custom_aggregation_data_raw, dict):
            print("DEBUG: Found 'custom_aggregation_results'. Converting string keys to tuples...")
            custom_aggregation_data_processed = {}
            custom_converted_count = 0
            custom_conversion_errors = 0
            # Reuse the same regex pattern
            decimal_pattern = re.compile(r"Decimal\('(-?\d*\.?\d+)'\)")

            for key_str, value_dict in custom_aggregation_data_raw.items():
                processed_key_str = key_str
                try:
                    processed_key_str = decimal_pattern.sub(r"'\1'", key_str)
                    key_tuple = ast.literal_eval(processed_key_str)

                    # Apply the same post-processing as standard aggregation if needed
                    # (Assuming the structure PO, Item, [Optional Price] is consistent)
                    final_key_list = []
                    if isinstance(key_tuple, tuple) and len(key_tuple) >= 2: # Custom might only have PO, Item
                        # PO Number (Index 0): Keep as string
                        final_key_list.append(str(key_tuple[0]))
                        # Item Number (Index 1): Keep as string
                        final_key_list.append(str(key_tuple[1]))
                        # Keep remaining elements (e.g., None in the example)
                        if len(key_tuple) > 2:
                            final_key_list.extend(key_tuple[2:])
                    else:
                        print(f"Warning: Custom key tuple '{key_tuple}' doesn't have expected length >= 2. Using original items.")
                        final_key_list = list(key_tuple)

                    final_key_tuple = tuple(final_key_list)

                    if isinstance(final_key_tuple, tuple):
                        custom_aggregation_data_processed[final_key_tuple] = value_dict
                        custom_converted_count += 1
                    else:
                        print(f"Warning: Final custom key is not a tuple for processed key string '{processed_key_str}'. Original: '{key_str}'. Result: {final_key_tuple}")
                        custom_conversion_errors += 1
                except (ValueError, SyntaxError, NameError, TypeError) as e:
                    print(f"Warning: Could not convert custom aggregation key string '{key_str}' (processed: '{processed_key_str}') to tuple: {e}")
                    custom_conversion_errors += 1

            invoice_data["custom_aggregation_results"] = custom_aggregation_data_processed
            print(f"DEBUG: Finished key conversion for custom_aggregation_results. Converted: {custom_converted_count}, Errors: {custom_conversion_errors}")
        # --- END CUSTOM AGGREGATION KEY CONVERSION ---

        return invoice_data
    except json.JSONDecodeError as e: print(f"Error: Invalid JSON in data file {data_path}: {e}"); return None
    except pickle.UnpicklingError as e: print(f"Error: Could not unpickle data file {data_path}: {e}"); return None
    except FileNotFoundError: print(f"Error: Data file not found at {data_path}"); return None
    except Exception as e: print(f"Error loading data file {data_path}: {e}"); traceback.print_exc(); return None
# --- End Placeholder ---

def calculate_header_dimensions(header_layout: List[Dict[str, Any]]) -> Tuple[int, int]:
    """
    Calculates the total number of rows and columns a header will occupy.

    Args:
        header_layout: The list of dictionaries defining the header.

    Returns:
        A tuple containing (num_header_rows, num_header_columns).
        Returns (0, 0) if the layout is empty.
    """
    if not header_layout:
        return (0, 0)

    # Calculate the total number of rows the header occupies
    num_rows = max(cell.get('row', 0) + cell.get('rowspan', 1) for cell in header_layout)

    # Calculate the total number of columns the header occupies
    num_cols = max(cell.get('col', 0) + cell.get('colspan', 1) for cell in header_layout)

    return (num_rows, num_cols)

# Add these to your imports at the top of the file
from openpyxl.worksheet.worksheet import Worksheet 
# You already have List, Dict, Any, Tuple from typing

def pre_calculate_and_insert_rows(
    worksheet: Worksheet,
    sheet_name: str,
    start_row: int,
    table_keys: List[str],
    all_tables_data: Dict[str, Any],
    sheet_mapping_section: Dict[str, Any],
    header_to_write: List[Dict[str, Any]]
) -> Tuple[bool, int]:
    """
    Pre-calculates the total number of rows required for a multi-table layout and inserts them.

    Args:
        worksheet: The openpyxl worksheet object to modify.
        sheet_name: The name of the sheet (for logging).
        start_row: The starting row index for insertion.
        table_keys: The sorted list of keys for the tables to be processed.
        all_tables_data: The dictionary containing all table data.
        sheet_mapping_section: The configuration section for the current sheet.
        header_to_write: The header layout definition.

    Returns:
        A tuple containing:
        - bool: True if the rows were inserted successfully, False otherwise.
        - int: The total number of rows that were calculated and inserted.
    """
    # --- Pre-calculation ---
    total_rows_to_insert = 0
    num_tables = len(table_keys)
    add_blank_after_hdr_flag = sheet_mapping_section.get("add_blank_after_header", False)
    add_blank_before_ftr_flag = sheet_mapping_section.get("add_blank_before_footer", False)
    final_row_spacing = sheet_mapping_section.get('row_spacing', 0)
    summary_flag = sheet_mapping_section.get("summary", False)

    print("--- Pre-calculating total rows for multi-table section ---")
    for i, table_key in enumerate(table_keys):
        table_data_to_fill = all_tables_data.get(str(table_key))
        if not table_data_to_fill or not isinstance(table_data_to_fill, dict):
            continue

        num_header_rows, _ = calculate_header_dimensions(header_to_write)
        total_rows_to_insert += num_header_rows
        print(f"  Table {table_key}: +{num_header_rows} (header)")

        if add_blank_after_hdr_flag:
            total_rows_to_insert += 1
            print(f"  Table {table_key}: +1 (blank after header)")

        max_len = max((len(v) for v in table_data_to_fill.values() if isinstance(v, list)), default=0)
        num_data_rows = max_len
        total_rows_to_insert += num_data_rows
        print(f"  Table {table_key}: +{num_data_rows} (data rows)")

        if add_blank_before_ftr_flag:
            total_rows_to_insert += 1
            print(f"  Table {table_key}: +1 (blank before footer)")

        total_rows_to_insert += 1
        print(f"  Table {table_key}: +1 (footer)")

        if i < num_tables - 1:
            total_rows_to_insert += 1
            print(f"  Table {table_key}: +1 (spacer)")

    if num_tables > 1:
        total_rows_to_insert += 1
        print("  Overall: +1 (Grand Total Row)")

    if summary_flag and num_tables > 0:
        total_rows_to_insert += 2
        print("  Overall: +2 (Summary Flag Rows)")

    if final_row_spacing > 0:
        total_rows_to_insert += final_row_spacing
        print(f"  Overall: +{final_row_spacing} (Final Spacing)")

    print(f"--- Total rows to insert for multi-table section: {total_rows_to_insert} ---")

    # --- Bulk Insert ---
    if total_rows_to_insert > 0:
        try:
            print(f"Inserting {total_rows_to_insert} rows at index {start_row} for sheet '{sheet_name}'...")
            worksheet.insert_rows(start_row, amount=total_rows_to_insert)
            invoice_utils.safe_unmerge_block(worksheet, start_row, start_row + total_rows_to_insert - 1, worksheet.max_column)
            print("Bulk rows inserted and unmerged successfully.")
            return True, total_rows_to_insert
        except Exception as bulk_insert_err:
            print(f"ERROR: Failed bulk row insert for multi-table: {bulk_insert_err}")
            return False, 0
    
    return True, 0 # Succeeded, but inserted 0 rows



# --- Main Orchestration Logic ---
def main():
    """Main function to orchestrate invoice generation."""
    parser = argparse.ArgumentParser(description="Generate Invoice from Template and Data using configuration files.")
    parser.add_argument("input_data_file", help="Path to the input data file (.json or .pkl). Filename base determines template/config.")
    parser.add_argument("-o", "--output", default="result.xlsx", help="Path for the output Excel file (default: result.xlsx)")
    parser.add_argument("-t", "--templatedir", default="./TEMPLATE", help="Directory containing template Excel files (default: ./TEMPLATE)")
    parser.add_argument("-c", "--configdir", default="./configs", help="Directory containing configuration JSON files (default: ./configs)")
    parser.add_argument("--fob", action="store_true", help="Generate FOB version using final_fob_compounded_result for Invoice/Contract sheets.")
    parser.add_argument("--custom", action="store_true", help="Enable custom processing logic (details TBD).")
    args = parser.parse_args()

    print("--- Starting Invoice Generation ---")
    print(f"Input Data: {args.input_data_file}"); print(f"Template Dir: {args.templatedir}"); print(f"Config Dir: {args.configdir}"); print(f"Output File: {args.output}")

    print("\n1. Deriving file paths..."); paths = derive_paths(args.input_data_file, args.templatedir, args.configdir)
    if not paths: sys.exit(1)

    print("\n2. Loading configuration and data..."); config = load_config(paths['config']); invoice_data = load_data(paths['data'])
    if not config or not invoice_data: sys.exit(1)

    print(f"\n3. Copying template '{paths['template'].name}' to '{args.output}'..."); output_path = Path(args.output).resolve()
    try:
        output_path.parent.mkdir(parents=True, exist_ok=True);
        shutil.copy(paths['template'], output_path);
    except Exception as e:
        print(f"Error copying template: {e}"); sys.exit(1)
    print(f"Template copied successfully to {output_path}")

    print("\n4. Processing workbook...");
    workbook = None; processing_successful = True

    try:
        workbook = openpyxl.load_workbook(output_path)

        # --- Determine sheets to process ---
        sheets_to_process_config = config.get('sheets_to_process', [])
        if not sheets_to_process_config:
            sheets_to_process = [workbook.active.title] if workbook.active else []
        else:
            sheets_to_process = [s for s in sheets_to_process_config if s in workbook.sheetnames] # Filter valid sheets

        if not sheets_to_process:
            print("Error: No valid sheets found or specified to process.")
            if workbook:
                try: workbook.close()
                except Exception: pass
            sys.exit(1) # Exit if no sheets to process

        # --- Store Original Merges BEFORE processing using merge_utils ---
        original_merges = merge_utils.store_original_merges(workbook, sheets_to_process)
        print("DEBUG: Stored original merges structure:")

        # --- Get other config sections ---
        sheet_data_map = config.get('sheet_data_map', {})
        global_footer_rules = config.get('footer_rules', {})
        data_mapping_config = config.get('data_mapping', {})

        # ***** MOVED GLOBAL PALLET CALCULATION HERE *****
        final_grand_total_pallets = 0
        print("DEBUG: Pre-calculating final grand total pallets globally...")
        processed_tables_data_for_calc = invoice_data.get('processed_tables_data', {})
        if isinstance(processed_tables_data_for_calc, dict) and processed_tables_data_for_calc:
            temp_total = 0
            table_keys_for_calc = sorted(processed_tables_data_for_calc.keys(), key=lambda x: int(x) if str(x).isdigit() else float('inf'))
            for temp_key in table_keys_for_calc:
                temp_table_data = processed_tables_data_for_calc.get(str(temp_key))
                if isinstance(temp_table_data, dict):
                    pallet_counts = temp_table_data.get("pallet_count", [])
                    if isinstance(pallet_counts, list):
                        for count in pallet_counts:
                            try:
                                temp_total += int(count)
                            except (ValueError, TypeError):
                                pass # Ignore non-integer counts
            final_grand_total_pallets = temp_total
        else:
            print("DEBUG: 'processed_tables_data' not found or empty in input data. final_grand_total_pallets remains 0.")
        print(f"DEBUG: Globally calculated final grand total pallets: {final_grand_total_pallets}")
        # ***** END GLOBAL PALLET CALCULATION *****


        print(f"\nWill process sheets: {sheets_to_process}")
        # --- Start Sheet Processing Loop ---
        for sheet_name in sheets_to_process:
            print(f"\n--- Processing Sheet: '{sheet_name}' ---")
            if sheet_name not in workbook.sheetnames:
                print(f"Warning: Sheet '{sheet_name}' not found at processing time. Skipping.")
                continue
            worksheet = workbook[sheet_name]

            # --- Get sheet-specific config sections ---
            sheet_mapping_section = data_mapping_config.get(sheet_name, {}) # Use .get for safety
            data_source_indicator = sheet_data_map.get(sheet_name) # Get indicator from config

            # --- Check for FOB flag override ---
            if args.fob and sheet_name in ["Invoice", "Contract"]:
                print(f"DEBUG: --fob flag active. Overriding data source for '{sheet_name}' to 'fob_aggregation'.")
                data_source_indicator = 'fob_aggregation'
            # --- End FOB flag override ---

            sheet_styling_config = sheet_mapping_section.get("styling") # Get styling rules dict or None

            if not sheet_mapping_section: print(f"Warning: No 'data_mapping' section for sheet '{sheet_name}'. Skipping."); continue
            if not data_source_indicator: print(f"Warning: No 'sheet_data_map' entry for sheet '{sheet_name}' (or FOB override failed). Skipping."); continue # Adjusted warning

            # --- Retrieve flags and mappings ONCE per sheet ---
            add_blank_after_hdr_flag = sheet_mapping_section.get("add_blank_after_header", False)
            static_content_after_hdr_dict = sheet_mapping_section.get("static_content_after_header", {})
            add_blank_before_ftr_flag = sheet_mapping_section.get("add_blank_before_footer", False)
            static_content_before_ftr_dict = sheet_mapping_section.get("static_content_before_footer", {})
            sheet_inner_mapping_rules_dict = sheet_mapping_section.get('mappings', {})
            final_row_spacing = sheet_mapping_section.get('row_spacing', 0)
            merge_rules_after_hdr = sheet_mapping_section.get("merge_rules_after_header", {})
            merge_rules_before_ftr = sheet_mapping_section.get("merge_rules_before_footer", {})
            merge_rules_footer = sheet_mapping_section.get("merge_rules_footer", {})
            data_cell_merging_rules = sheet_mapping_section.get("data_cell_merging_rule", None)
            sheet_header_to_write = sheet_mapping_section.get("header_to_write", None)

            print(f"DEBUG Check Flags Read for Sheet '{sheet_name}': after_hdr={add_blank_after_hdr_flag}, before_ftr={add_blank_before_ftr_flag}")
            if sheet_styling_config: print("DEBUG: Styling config found for this sheet.")
            else: print("DEBUG: No styling config found for this sheet.")

            all_tables_data = invoice_data.get('processed_tables_data', {})
            table_keys = sorted(all_tables_data.keys(), key=lambda x: int(x) if str(x).isdigit() else float('inf'))
            # ================================================================
            # --- Handle Multi-Table Case (e.g., Packing List) ---
            # ================================================================
            if data_source_indicator == "processed_tables_multi":
                print(f"Processing sheet '{sheet_name}' as multi-table (write header mode).")
                all_tables_data = invoice_data.get('processed_tables_data', {})
                if not all_tables_data or not isinstance(all_tables_data, dict): print(f"Warning: 'processed_tables_data' not found/valid. Skipping '{sheet_name}'."); continue

                header_to_write = sheet_mapping_section.get('header_to_write');
                header_merge_rules = sheet_mapping_section.get('header_merge_rules')
                start_row = sheet_mapping_section.get('start_row') # Use config start_row
                if not start_row or not header_to_write: print(f"Error: Config for multi-table '{sheet_name}' missing 'start_row' or 'header_to_write'. Skipping."); processing_successful = False; continue

                table_keys = sorted(all_tables_data.keys(), key=lambda x: int(x) if str(x).isdigit() else float('inf'))
                print(f"Found table keys in data: {table_keys}"); num_tables = len(table_keys); last_table_header_info = None

                # --- Call the new refactored function ---
                success, _ = pre_calculate_and_insert_rows(
                    worksheet=worksheet,
                    sheet_name=sheet_name,
                    start_row=start_row,
                    table_keys=table_keys,
                    all_tables_data=all_tables_data,
                    sheet_mapping_section=sheet_mapping_section,
                    header_to_write=header_to_write
                )
                spacer_row = 1

                if not success:
                    processing_successful = False
                    continue # Skip to the next sheet if insertion failed
                # --- End function call ---

                # --- V11: Initialize write pointer --- 
                write_pointer_row = start_row # Start writing at the beginning of the inserted block
 
                # ***** INITIALIZE GRAND TOTAL FOR *THIS SHEET TYPE* & PALLET ORDER VARIABLES *****
                grand_total_pallets_for_summary_row = 0
                all_data_ranges = [] # List to store tuples of (start_row, end_row) for SUM
                # ***** END INITIALIZE *****
 
                # ***** REMOVED REDUNDANT PRE-CALCULATION LOOP - NOW DONE GLOBALLY *****
 
                # --- V11: Main loop now only writes data, doesn't insert --- # TODO urgent
                for i, table_key in enumerate(table_keys):
                    print(f"\nProcessing table key: '{table_key}' ({i+1}/{num_tables})")
                    table_data_to_fill = all_tables_data.get(str(table_key))
                    if not table_data_to_fill or not isinstance(table_data_to_fill, dict): print(f"Warning: No/invalid data for table key '{table_key}'. Skipping."); continue
 
                    print(f"Writing header for table '{table_key}' at row {write_pointer_row}...");
                    written_header_info = invoice_utils.write_header(
                        worksheet, write_pointer_row, sheet_header_to_write, sheet_styling_config
                    )
                    if not written_header_info: print(f"Error writing header for table '{table_key}'. Skipping sheet."); processing_successful = False; break
                    last_table_header_info = written_header_info # Keep track for width setting later
 
                    # Update write pointer after header
                    num_header_rows, num_columns = calculate_header_dimensions(sheet_header_to_write)
                    write_pointer_row += num_header_rows
 
                    print(f"Filling data and footer for table '{table_key}' starting near row {write_pointer_row}...")
                    # Pass the current write pointer as the effective 'start row' for fill_invoice_data
                    # It will write header, data, footer starting from here
                    # NOTE: We need to adjust fill_invoice_data to use the passed start row correctly
                    #       instead of header_info['second_row_index'] + 1
 
                    # Modify the header_info dict passed to fill_invoice_data dynamically
                    temp_header_info = written_header_info.copy()
                    temp_header_info['first_row_index'] = write_pointer_row - num_header_rows # The row wherea header started
                    temp_header_info['second_row_index'] = temp_header_info['first_row_index'] + 1 # The last row of the header
 
                    fill_success, next_row_after_chunk, data_start, data_end, table_pallets = invoice_utils.fill_invoice_data(
                        worksheet=worksheet,
                        sheet_name=sheet_name,
                        sheet_config=sheet_mapping_section, # Pass current sheet's config
                        all_sheet_configs=data_mapping_config, # <--- Pass the full config map
                        data_source=table_data_to_fill,
                        data_source_type='processed_tables',
                        header_info=temp_header_info,
                        mapping_rules=sheet_inner_mapping_rules_dict,
                        sheet_styling_config=sheet_styling_config,
                        add_blank_after_header=add_blank_after_hdr_flag,
                        static_content_after_header=static_content_after_hdr_dict,
                        add_blank_before_footer=add_blank_before_ftr_flag,
                        static_content_before_footer=static_content_before_ftr_dict,
                        merge_rules_after_header=merge_rules_after_hdr,
                        merge_rules_before_footer=merge_rules_before_ftr,
                        merge_rules_footer=merge_rules_footer,
                        footer_info=None, max_rows_to_fill=None,
                        grand_total_pallets=final_grand_total_pallets,
                        custom_flag=args.custom,
                        data_cell_merging_rules=data_cell_merging_rules,
                    )
                    # fill_invoice_data now handles writing blank rows, data, footer row
                    # within the allocated space. next_row_after_chunk is the row AFTER its footer.
 
                    if fill_success:
                        num_cols_spacer = 1
                        print(f"Finished table '{table_key}'. Next available write pointer is {next_row_after_chunk}")
                        grand_total_pallets_for_summary_row += table_pallets
                        if data_start > 0 and data_end >= data_start: all_data_ranges.append((data_start, data_end))
 
                        write_pointer_row = next_row_after_chunk # Update pointer to be after the chunk
 
                        is_last_table = (i == num_tables - 1)
                        if not is_last_table: 
                            # Write the spacer row content (optional, could just be blank)
                            spacer_row = write_pointer_row
                            if num_cols_spacer > 0:
                                try:
                                    print(f"Writing merged spacer row at {spacer_row} across {num_cols_spacer} columns...")
                                    # No insert needed, just merge and maybe clear/style
                                    invoice_utils.unmerge_row(worksheet, spacer_row, num_cols_spacer) # Ensure clear
                                    worksheet.merge_cells(start_row=spacer_row, start_column=1, end_row=spacer_row, end_column=num_cols_spacer)
                                    # Optionally add styling or blank value to the merged cell
                                    # worksheet.cell(row=spacer_row, column=1).border = ... 
                                    write_pointer_row += 1 # Advance pointer past the spacer row
                                except Exception as merge_err: 
                                    print(f"Warning: Failed to write/merge spacer row {spacer_row}: {merge_err}"); 
                                    write_pointer_row += 1 # Still advance pointer even if merge fails
                            else: 
                                print("Warning: Cannot determine table width for spacer.");
                                write_pointer_row += 1 # Advance pointer anyway
                        # No 'else' needed, pointer is already correct if it's the last table
                    else: 
                        print(f"Error filling data/footer for table '{table_key}'. Stopping."); 
                        processing_successful = False; break
                # --- End Table Loop ---
 
                # ***** ADD GRAND TOTAL ROW (for multi-table summary) *****
                    if processing_successful and num_tables > 1:
                        grand_total_row_num = write_pointer_row
                        print(f"\n--- Adding Grand Total Row at index {grand_total_row_num} using write_footer_row ---")
                        try:
                            # Get the footer configuration from the sheet's mapping section
                            footer_config_for_gt = sheet_mapping_section.get("footer_configurations", {})

                            # Call the reusable write_footer_row function with the correct arguments
                            footer_row_index = invoice_utils.write_footer_row(
                                worksheet=worksheet,
                                footer_row_num=grand_total_row_num,
                                header_info=last_table_header_info,
                                sum_ranges=all_data_ranges,
                                footer_config=footer_config_for_gt,
                                pallet_count=grand_total_pallets_for_summary_row,
                                override_total_text="TOTAL OF:",
                                grand_total_flag=True
                            )

                            if footer_row_index != -1:
                                write_pointer_row += 1 # Advance pointer after the new row
                                print(f"--- Finished Adding Grand Total Row. Next write pointer: {write_pointer_row} ---")
                            else:
                                print("--- ERROR: write_footer_row failed to generate the Grand Total row. ---")
                                processing_successful = False

                        except Exception as gt_err:
                            print(f"--- ERROR preparing for or calling write_footer_row for Grand Total: {gt_err} ---")
                            traceback.print_exc()
                    # ***** END REVISED GRAND TOTAL ROW *****
                # --- V11: Logic for Summary Rows (BUFFALO summary + blank) ---
                summary_flag = sheet_mapping_section.get("summary", False)
                sheet_inner_mapping_rules_dict = sheet_mapping_section.get('mappings', {})
                if summary_flag and processing_successful and last_table_header_info:
                    # Get the footer config to pass its styles to the summary writer
                    footer_config_for_summary = sheet_mapping_section.get("footer_configurations", {})
                    
                    write_pointer_row = invoice_utils.write_summary_rows(
                        worksheet=worksheet,
                        start_row=write_pointer_row,
                        header_info=last_table_header_info,
                        all_tables_data=all_tables_data,
                        table_keys=table_keys,
                        footer_config=footer_config_for_summary, # <-- Pass the config here
                        mapping_rules=sheet_inner_mapping_rules_dict,
                        styling_config=sheet_styling_config
                    )
                # --- End Summary Rows Logic ---
                # --- START FOB-Specific Hardcoded Replacements (Multi-Table) ---
                if args.fob:
                    # Apply to the current multi-table sheet if --fob is active
                    print(f":::::::::::::::::::: Performing FOB-specific hardcoded replacements for multi-table sheet '{sheet_name}'...")
                    fob_specific_replacement_rules = [
                        {
                            "find": "DAP",          # Text to find
                            "replace": "FOB",       # Text to replace with
                            "case_sensitive": False # Match regardless of case
                        },
                        {
                            "find": "FCA",          # Text to find
                            "replace": "FOB",       # Text to replace with
                            "case_sensitive": False # Match regardless of case
                        },
                        {
                            "find": "BAVET, SVAY RIENG",    # Text to find
                            "replace": "BAVET",     # Text to replace with
                            "case_sensitive": True, # Match regardless of case
                            "exact_cell_match": True
                        },
                        {
                            "find": "BINH PHUOC",    # Text to find
                            "replace": "BAVET",     # Text to replace with
                            "case_sensitive": True, # Match regardless of case
                            "exact_cell_match": True
                        }
                        # Add more FOB-specific rules here if needed
                    ]
                    # Call the function designed for find/replace rules
                    # Apply ONLY to the current sheet being processed
                    text_replace_utils.find_and_replace_in_workbook(
                        workbook=workbook,
                        replacement_rules=fob_specific_replacement_rules,
                        target_sheets=[sheet_name] # Apply only to the current sheet
                    )
                    print(f":::::::::::::::::::: Finished FOB-specific hardcoded replacements for multi-table sheet '{sheet_name}'.")
                # --- END FOB-Specific Hardcoded Replacements (Multi-Table) ---

                # --- Apply Column Widths AFTER loop using the last header info ---
                if processing_successful and last_table_header_info:
                    print(f"Applying column widths for multi-table sheet '{sheet_name}'...")
                    invoice_utils.apply_column_widths(
                        worksheet,
                        sheet_styling_config,
                        last_table_header_info.get('column_map')
                    )
                # --- End Apply Column Widths ---
 
                # --- Final Spacer Rows --- 
                # V11: No insert needed, just advance pointer if required
                if final_row_spacing > 0 and num_tables > 0 and processing_successful:
                    final_spacer_start_row = write_pointer_row
                    try:
                        print(f"Config requests final spacing ({final_row_spacing}). Advancing pointer from {final_spacer_start_row}.")
                        # worksheet.insert_rows(final_spacer_start_row, amount=final_row_spacing) # REMOVED INSERT
                        # Optionally clear/style these rows
                        write_pointer_row += final_row_spacing
                        print(f"Pointer advanced for final spacing. Final pointer: {write_pointer_row}")
                    except Exception as final_spacer_err: 
                        print(f"Warning: Error during final spacing logic: {final_spacer_err}")
 
            # ================================================================
            # --- Handle Single Table / Aggregation Case ---
            # ================================================================
            else:
                print(f"Processing sheet '{sheet_name}' as single table/aggregation.")
                header_info = None; footer_info = None

                start_row = sheet_mapping_section.get('start_row');
                header_to_write = sheet_mapping_section.get('header_to_write');
                if not start_row or not header_to_write: print(f"Error: Config for '{sheet_name}' missing 'start_row' or 'header_to_write'. Skipping."); processing_successful = False; continue

                print(f"Writing header at row {start_row}...");
                header_info = invoice_utils.write_header(
                    worksheet, start_row, header_to_write, sheet_styling_config
                )
                if not header_info: print(f"Error: Failed to write header for '{sheet_name}'. Skipping."); processing_successful = False; continue
                print(f"DEBUG: Header Info for '{sheet_name}':")
                print(f"  - Column Map: {header_info.get('column_map')}")
                print(f"Header written successfully.")

                # --- Get Data Source ---
                data_to_fill = None; data_source_type = None
                print(f"DEBUG: Retrieving data source for '{sheet_name}' using indicator: '{data_source_indicator}'")

                # --- Custom Flag Logic ---
                if args.custom and data_source_indicator == 'aggregation':
                    print(f"DEBUG: --custom flag active. Attempting to use 'custom_aggregation_results' for sheet '{sheet_name}'.")
                    data_to_fill = invoice_data.get('custom_aggregation_results')
                    data_source_type = 'aggregation' # Type remains aggregation
                    if data_to_fill is not None:
                        print("DEBUG: Successfully retrieved 'custom_aggregation_results'.")
                    else:
                        print("Warning: 'custom_aggregation_results' key not found in data, despite --custom flag.")
                # --- End Custom Flag Logic ---

                # --- Standard Data Source Selection (if not overridden by --custom) ---
                if data_to_fill is None: # Only proceed if custom logic didn't already set it
                    if data_source_indicator == 'fob_aggregation': # Check for FOB first
                        data_to_fill = invoice_data.get('final_fob_compounded_result'); data_source_type = 'fob_aggregation';
                        print("DEBUG: Using 'final_fob_compounded_result' data (FOB mode).")
                        
                        # Add mapping for combined_description to Description column
                        if isinstance(data_to_fill, dict):
                            for map_key, map_rule in sheet_inner_mapping_rules_dict.items():
                                if isinstance(map_rule, dict):
                                    header = map_rule.get('header')
                                    if header in ['Description', 'DESCRIPTION OF GOODS', 'Description of Goods']:
                                        # Override any existing mapping to use combined_description
                                        map_rule['value_key'] = 'combined_description'
                                        print(f"DEBUG: FOB Mode - Mapped combined_description to {header}")
                    elif data_source_indicator == 'aggregation': # Standard aggregation
                        data_to_fill = invoice_data.get('standard_aggregation_results'); data_source_type = 'aggregation';
                        print("DEBUG: Using 'standard_aggregation_results' data.")
                    elif 'processed_tables_data' in invoice_data and data_source_indicator in invoice_data.get('processed_tables_data', {}):
                        data_to_fill = invoice_data['processed_tables_data'].get(data_source_indicator); data_source_type = 'processed_tables';
                        print(f"DEBUG: Using 'processed_tables_data' key '{data_source_indicator}' (Likely for multi-table).")
                    else:
                        print(f"DEBUG: Data source indicator '{data_source_indicator}' not found in expected locations ('final_fob_compounded_result', 'standard_aggregation_results', 'custom_aggregation_results' or 'processed_tables_data').")
                # --- End Standard Data Source Selection ---

                if data_to_fill is None: print(f"Warning: Data source '{data_source_indicator}' unknown or data empty. Skipping fill."); continue

                # *** Explicitly re-check header_info before the call ***
                # header_info = written_header_info # Re-assign just in case (Use the variable name from the write_header call if different)
                # Let's assume the variable was indeed header_info
                if header_info is None:
                    print(f"FATAL DEBUG: header_info is unexpectedly None right before calling fill_invoice_data for {sheet_name}")

                if header_info is not None and header_info.get('column_map'):
                    print(f"DEBUG: Calling fill_invoice_data for single table '{sheet_name}'")
                    print(f"DEBUG: header_info passed to fill_invoice_data: {header_info}") # <--- SHOWS None HERE
                    # *** Get aggregated FOB values ---
                    total_sqft = data_to_fill.get('total_sqft', 0)
                    total_amount = data_to_fill.get('total_amount', 0)
                    combined_po = data_to_fill.get('combined_po')
                    combined_item = data_to_fill.get('combined_item')
                    combined_description = data_to_fill.get('combined_description')

                    # *** Map headers to data source keys and find target columns ---
                    fob_data_keys = {
                        'P.O N °': 'combined_po', 'P.O Nº': 'combined_po',
                        'ITEM NO': 'combined_item', 'ITEM Nº': 'combined_item',
                        'Quantity ( SF )': 'total_sqft', 'Quantity(SF)': 'total_sqft',
                        'Amount ( USD )': 'total_amount', 'Total value(USD)': 'total_amount',
                        'Description': 'combined_description', 'DESCRIPTION OF GOODS': 'combined_description',
                        'Description of Goods': 'combined_description', "Description Of Goods": 'combined_item'
                    }

                    # ***** MODIFIED CALL for single table (Removed starting_pallet_order, adjusted return) *****
                    fill_success, next_row_after_footer, dog, _, table_pallets = invoice_utils.fill_invoice_data(
                        worksheet=worksheet,
                        sheet_name=sheet_name,
                        sheet_config=sheet_mapping_section, # Pass current sheet's config
                        all_sheet_configs=data_mapping_config, # <--- Pass the full config map
                        data_source=data_to_fill,
                        data_source_type=data_source_type,
                        header_info=header_info,
                        mapping_rules=sheet_inner_mapping_rules_dict,
                        sheet_styling_config=sheet_styling_config,
                        add_blank_after_header=add_blank_after_hdr_flag,
                        static_content_after_header=static_content_after_hdr_dict,
                        add_blank_before_footer=add_blank_before_ftr_flag,
                        static_content_before_footer=static_content_before_ftr_dict,
                        merge_rules_after_header=merge_rules_after_hdr,
                        merge_rules_before_footer=merge_rules_before_ftr,
                        merge_rules_footer=merge_rules_footer,
                        footer_info=footer_info, max_rows_to_fill=None,
                        grand_total_pallets=final_grand_total_pallets,
                        custom_flag=args.custom,
                        data_cell_merging_rules=data_cell_merging_rules
                    )
                    # ***** END MODIFIED CALL *****
                    
                    # Initialize sheet_grand_totals dictionary
                    sheet_grand_totals = {
                        "grand_total_nett_weight": 0.0,
                        "grand_total_gross_weight": 0.0,
                        "grand_total_cbm": 0.0,  # <-- STEP 1: ADDED NEW KEY FOR CBM
                        # Add other totals needed for rows_after_footer here
                    }

                    # This block calculates totals if the sheet is a multi-table sheet (e.g. Packing List)
                    # and these totals are then used by write_configured_rows.
                    # The condition `processing_successful and all_tables_data` applies.
                    # `all_tables_data` is typically `invoice_data.get('processed_tables_data', {})`
                    # `table_keys` are the sorted keys from `all_tables_data`
                    if processing_successful and all_tables_data: # This condition might be more relevant for multi-table sheets
                        print("Calculating totals for rows_after_footer...")
                        for table_key in table_keys: # table_keys would be derived from all_tables_data
                            table_data = all_tables_data.get(str(table_key))
                            if table_data and isinstance(table_data, dict):
                                try:
                                    # Summing 'net' values - ensure 'net' exists and contains numbers
                                    nett_weights = table_data.get("net", [])
                                    if isinstance(nett_weights, list):
                                        sheet_grand_totals["grand_total_nett_weight"] += sum(
                                            float(nw) for nw in nett_weights 
                                            if isinstance(nw, (int, float, str)) and str(nw).strip() # Check if convertible
                                        )
                                except (ValueError, TypeError) as e:
                                    print(f"Warning: Error summing nett weights for table {table_key}: {e}")
                                try:
                                    # Summing 'gross' values - ensure 'gross' exists and contains numbers
                                    gross_weights = table_data.get("gross", [])
                                    if isinstance(gross_weights, list):
                                        sheet_grand_totals["grand_total_gross_weight"] += sum(
                                            float(gw) for gw in gross_weights 
                                            if isinstance(gw, (int, float, str)) and str(gw).strip() # Check if convertible
                                        )
                                except (ValueError, TypeError) as e:
                                    print(f"Warning: Error summing gross weights for table {table_key}: {e}")
                                
                                # --- STEP 2: ADDED CALCULATION LOGIC FOR CBM ---
                                try:
                                    # Summing 'cbm' values - ensure 'cbm' exists and contains numbers
                                    cbm_values = table_data.get("cbm", []) # Assuming 'cbm' is the key in your data
                                    if isinstance(cbm_values, list):
                                        sheet_grand_totals["grand_total_cbm"] += sum(
                                            float(cbm_val) for cbm_val in cbm_values 
                                            if isinstance(cbm_val, (int, float, str)) and str(cbm_val).strip() # Check if convertible
                                        )
                                except (ValueError, TypeError) as e:
                                    print(f"Warning: Error summing CBM for table {table_key}: {e}")
                                # --- END STEP 2 ---

                        print(f"Calculated Sheet Grand Totals: {sheet_grand_totals}")
                        
                        # This part calls write_configured_rows. 
                        # Ensure last_table_header_info is relevant or use header_info for single table sheets.
                        # For single table sheets, header_info should be used.
                        # The write_pointer_row also needs to be correctly set. For single table, it would be next_row_after_footer.
                        
                        current_header_info_for_rows_after_footer = header_info # For single table sheets
                        current_write_pointer_for_rows_after_footer = next_row_after_footer

                        # If this logic is inside the "processed_tables_multi" part of your script, 
                        # then last_table_header_info and write_pointer_row (as updated by the multi-table loop) would be correct.
                        # The provided snippet seems to be from the single-table processing path due to `next_row_after_footer`.

                        if processing_successful and current_header_info_for_rows_after_footer:
                            rows_after_footer_enabled = sheet_mapping_section.get("rows_after_footer_enabled", False)
                            rows_after_footer_config = sheet_mapping_section.get("rows_after_footer", [])
                            
                            if rows_after_footer_enabled and isinstance(rows_after_footer_config, list) and len(rows_after_footer_config) > 0:
                                num_extra_rows = len(rows_after_footer_config)
                                
                                # For single-table sheets, insert rows if they are not already accounted for.
                                # This depends on whether fill_invoice_data already inserted space for these.
                                # Typically, fill_invoice_data inserts rows up to its own footer.
                                # So, rows *after* that footer need to be inserted if not already planned.
                                # However, write_configured_rows assumes rows are ALREADY INSERTED.
                                # So, you might need to insert rows *before* calling write_configured_rows.

                                print(f"DEBUG: Preparing to write {num_extra_rows} rows after footer, starting at potential row {current_write_pointer_for_rows_after_footer}")
                                print(f"DEBUG: Using header_info: {current_header_info_for_rows_after_footer.get('num_columns')} columns")
                                print(f"DEBUG: Calculated totals for these rows: {sheet_grand_totals}")
                                if num_extra_rows > 0: 
                                    print(f"DEBUG: Inserting {num_extra_rows} rows at {current_write_pointer_for_rows_after_footer} for 'rows_after_footer'.")
                                    worksheet.insert_rows(current_write_pointer_for_rows_after_footer, amount=num_extra_rows)
                                    # Unmerge the newly inserted block to prevent styling/merge issues
                                    invoice_utils.safe_unmerge_block(
                                        worksheet, 
                                        current_write_pointer_for_rows_after_footer, 
                                        current_write_pointer_for_rows_after_footer + num_extra_rows - 1, 
                                        current_header_info_for_rows_after_footer['num_columns']
                                    )
                                    print(f"DEBUG: Inserted and unmerged {num_extra_rows} rows.")
                                else:
                                    print("DEBUG: No extra rows configured in 'rows_after_footer', no insertion needed.")

                                invoice_utils.write_configured_rows( 
                                    worksheet=worksheet,
                                    start_row_index=current_write_pointer_for_rows_after_footer, # This should be the row where these new lines start
                                    num_columns=current_header_info_for_rows_after_footer['num_columns'],
                                    rows_config_list=rows_after_footer_config,
                                    calculated_totals=sheet_grand_totals, 
                                    default_style_config=sheet_styling_config
                                )
                                next_row_after_footer = next_row_after_footer + 2

                                # Update the pointer if you were managing it manually for subsequent operations on this sheet
                                # current_write_pointer_for_rows_after_footer += num_extra_rows 
                                # next_row_after_footer = current_write_pointer_for_rows_after_footer # If this is the very end of sheet processing

                                print(f"DEBUG: Finished writing rows after footer. The next available row would conceptually be after these {num_extra_rows} rows.")
                    
                    


                    if fill_success:
                        print(f"Successfully filled table data/footer for sheet '{sheet_name}'.")

                        # --- Apply Column Widths AFTER filling ---
                        print(f"Applying column widths for sheet '{sheet_name}'...")
                        invoice_utils.apply_column_widths(
                            worksheet,
                            sheet_styling_config,
                            header_info.get('column_map')
                        )
                        # --- End Apply Column Widths ---

                        # --- Final Spacer Rows --- (Existing logic)
                        if final_row_spacing >= 1:
                            final_spacer_start_row = next_row_after_footer
                            try:
                                print(f"Config requests final spacing ({final_row_spacing}). Adding blank row(s) at {final_spacer_start_row}.")
                                worksheet.insert_rows(final_spacer_start_row, amount=final_row_spacing)
                            except Exception as final_spacer_err: print(f"Warning: Failed to insert final spacer rows: {final_spacer_err}")

                        # --- Fill Summary Fields --- (Existing logic)
                        print("Attempting to fill summary fields...")
                        summary_data_source = invoice_data.get('final_fob_compounded_result', {})
                        if not summary_data_source: print("Warning: 'final_fob_compounded_result' not found.")
                        elif not sheet_inner_mapping_rules_dict: print("Warning: 'mappings' dict not found for summary fields.")
                        else:
                            summary_fields_found = 0
                            for map_key, map_rule in sheet_inner_mapping_rules_dict.items():
                                if isinstance(map_rule, dict) and 'marker' in map_rule:
                                    marker_text = map_rule['marker']; summary_value = summary_data_source.get(map_key)
                                    if not marker_text: print(f"Warning: Marker text missing for summary key '{map_key}'."); continue
                                    if summary_value is None: print(f"Warning: Summary value for key '{map_key}' not found."); continue
                                    target_cell = invoice_utils.find_cell_by_marker(worksheet, marker_text)
                                    if target_cell:
                                        summary_fields_found += 1
                                        try: # Format as number if possible
                                            summary_value_num = float(summary_value);
                                            target_cell.value = summary_value_num; target_cell.number_format = '#,##0.00' # Default summary format
                                            print(f"Wrote summary {summary_value_num:.2f} to {target_cell.coordinate} ('{marker_text}')")
                                        except (ValueError, TypeError): # Fallback to string
                                            target_cell.value = str(summary_value)
                                            print(f"Wrote summary '{summary_value}' as string to {target_cell.coordinate} ('{marker_text}')")
                            if summary_fields_found == 0: print("No summary field markers found/matched.")
                            else: print(f"Processed {summary_fields_found} summary field(s).")

                        # --- Define and Apply Data-Driven Replacements HERE (Inside single-table loop) ---
                        print("Performing data-driven replacements for single-table sheet...")
                        # Define rules here: placeholder text, path to data within invoice_data, target sheets
                        # NOTE: Target sheets here might be redundant if we only apply to the *current* sheet
                        # Consider refining rules if replacement should only apply to the sheet being processed.
                        data_driven_replacement_rules = [
                            {
                                "find": "JFINV",
                                "data_path": ["processed_tables_data", "1", "inv_no", 0], # Path within invoice_data
                                "target_sheets": ["Invoice", "Contract", "Packing list"], # Apply ONLY to current sheet
                                "case_sensitive": True
                            },
                            {
                                "find": "JFTIME",
                                "data_path": ["processed_tables_data", "1", "inv_date", 0], # Corrected path based on JF.json analysis
                                "target_sheets": ["Invoice", "Contract", "Packing list"], # Apply ONLY to current sheet
                                "case_sensitive": True,
                                "is_date": True  # Mark this as a date field for special handling
                            },
                            {
                                "find": "JFREF",
                                # IMPORTANT: Determine the CORRECT key for reference from JF.json
                                # Using 'reference_code' as a placeholder - CHANGE IF WRONG
                                "data_path": ["processed_tables_data", "1", "inv_ref", 0],
                                "target_sheets": ["Invoice", "Contract", "Packing list"], # Apply ONLY to current sheet
                                "case_sensitive": True
                            },
                            # Add more rules as needed
                        ]
                        # Call the utility function
                        # Add the :::::::::::::::::::: debug print here
                        proc_tables_rep = invoice_data.get('processed_tables_data', {})
                        table_1_data_rep = proc_tables_rep.get('1', {})

                        text_replace_utils.process_data_driven_replacements(
                            workbook,
                            invoice_data,
                            data_driven_replacement_rules
                        )
                        # --- End Data-Driven Replacements for single-table ---

                        # --- START FOB-Specific Hardcoded Replacements ---
                        if args.fob:
                            # Apply to ALL sheets when --fob is active (but only runs when processing that specific sheet)
                            print(f":::::::::::::::::::: Performing FOB-specific hardcoded replacements for sheet '{sheet_name}'...")
                            fob_specific_replacement_rules = [
                                {
                                    "find": "DAP",          # Text to find
                                    "replace": "FOB",       # Text to replace with
                                    "case_sensitive": False # Match regardless of case
                                },
                                {
                                    "find": "FCA",          # Text to find
                                    "replace": "FOB",       # Text to replace with
                                    "case_sensitive": False # Match regardless of case
                                },
                                {
                                    "find": "BINH PHUOC",    # Text to find
                                    "replace": "BAVET",     # Text to replace with
                                    "case_sensitive": False, # Match regardless of case
                                    "exact_cell_match": True
                                },
                                {
                                    "find": "BAVET, SVAY RIENG",    # Text to find
                                    "replace": "BAVET",     # Text to replace with
                                    "case_sensitive": True, # Match regardless of case
                                    "exact_cell_match": True
                                }
                                # Add more FOB-specific rules here if needed
                            ]
                            # Call the function designed for find/replace rules
                            # The function will apply the rules only to the sheet specified in target_sheets
                            text_replace_utils.find_and_replace_in_workbook(
                                workbook=workbook,
                                replacement_rules=fob_specific_replacement_rules,
                                target_sheets=[sheet_name] # Apply only to the current sheet being processed
                            )
                            print(f":::::::::::::::::::: Finished FOB-specific hardcoded replacements for sheet '{sheet_name}'.")
                        # --- END FOB-Specific Hardcoded Replacements ---

                    else: print(f"Failed to fill table data/footer for sheet '{sheet_name}'."); processing_successful = False
                else:
                    print(f"Error: Cannot fill data for '{sheet_name}' because header_info or column_map is missing/invalid.")
                    processing_successful = False; continue # Skip filling if header info is bad

        # --- Restore Original Merges AFTER processing all sheets using merge_utils ---
        merge_utils.find_and_restore_merges_heuristic(workbook, original_merges, sheets_to_process)

        # 5. Save the final workbook
        print("\n--------------------------------")
        if processing_successful:
            print("5. Saving final workbook...")
            workbook.save(output_path); print(f"--- Workbook saved successfully: '{output_path}' ---")
        else:
            print("--- Processing completed with errors. Saving workbook (may be incomplete). ---")
            try:
                # Corrected the closing quote below
                workbook.save(output_path); print(f"--- Incomplete workbook saved to: '{output_path}' ---")
            except Exception as save_err:
                print(f"--- CRITICAL ERROR: Failed to save incomplete workbook: {save_err} ---")

    except Exception as e:
        print(f"\n--- UNHANDLED ERROR during workbook processing: {e} ---"); traceback.print_exc()
        if workbook and output_path: # Try to save error state
             try:
                 error_filename = output_path.stem + "_ERROR" + output_path.suffix; error_path = output_path.with_name(error_filename)
                 print(f"Attempting to save workbook state to {error_path}..."); workbook.save(error_path); print("Workbook state saved.")
             except Exception as final_save_err: print(f"--- Could not save workbook state after error: {final_save_err} ---")
    finally:
        if workbook:
            try: workbook.close(); print("Workbook closed.")
            except Exception: pass

    print("\n--- Invoice Generation Finished ---")

# --- Run Main ---
if __name__ == "__main__":
    main()