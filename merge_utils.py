import openpyxl
import traceback
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import range_boundaries, get_column_letter, column_index_from_string
# from openpyxl.worksheet.dimensions import RowDimension # Not strictly needed for access
from typing import Dict, List, Optional, Tuple, Any

# --- store_original_merges FILTERED to ignore merges ABOVE row 16 ---
def store_original_merges(workbook: openpyxl.Workbook, sheet_names: List[str]) -> Dict[str, List[Tuple[int, Any, Optional[float]]]]:
    """
    Stores the HORIZONTAL span (colspan), the value of the top-left cell,
    and the height of the starting row for merged ranges in specified sheets,
    ASSUMING all merges are only 1 row high AND **start at row 16 or below**.
    Merges starting above row 16 (row < 16) are ignored.
    WARNING: Does NOT store starting coordinates... (rest of docstring unchanged)

    Args: (args unchanged)

    Returns:
        A dictionary where keys are sheet names and values are lists of
        tuples: (col_span, top_left_cell_value, row_height).
        row_height will be None if the original row had default height.
    """
    original_merges = {}
    print("\nStoring original merge horizontal spans, top-left values, and row heights (NO coordinates)...")
    print("  (Ignoring merges that start above row 16)") # Updated filter info
    for sheet_name in sheet_names:
        if sheet_name in workbook.sheetnames:
            worksheet: Worksheet = workbook[sheet_name] # Type hint for clarity
            merges_data = []
            merged_ranges_copy = list(worksheet.merged_cells.ranges)
            skipped_above_16_count = 0 # Counter for this filter

            for merged_range in merged_ranges_copy:
                min_col, min_row, max_col, max_row = merged_range.bounds

                # --- Check 1: Skip if multi-row ---
                if max_row != min_row:
                    # print(f"  Skipping merge {merged_range.coord} on sheet '{sheet_name}' - it spans multiple rows ({min_row} to {max_row}).")
                    continue

                # ***** NEW CHECK 2: Skip if merge starts ABOVE row 16 *****
                if min_row < 16:
                    # print(f"  Skipping merge {merged_range.coord} on sheet '{sheet_name}' - starts at row {min_row} (above row 16).") # Keep commented unless needed
                    skipped_above_16_count += 1
                    continue
                # ***** END NEW CHECK *****

                # --- If not skipped, proceed to get span, height, value ---
                col_span = max_col - min_col + 1
                row_height = None # Default to None
                try:
                    # Get Row Height
                    row_dim = worksheet.row_dimensions[min_row]
                    row_height = row_dim.height
                    # print(f"    DEBUG Store: Sheet='{sheet_name}', MergeCoord='{merged_range.coord}', StartRow={min_row}, Storing Height={row_height} (Type: {type(row_height)})")

                    # Get Value
                    top_left_value = worksheet.cell(row=min_row, column=min_col).value

                    # Store Data (span, value, height)
                    merges_data.append((col_span, top_left_value, row_height))

                except KeyError:
                     print(f"    Warning: Could not find row dimension for row {min_row} on sheet '{sheet_name}' while getting height. Storing height as None.")
                     try:
                         top_left_value = worksheet.cell(row=min_row, column=min_col).value
                     except Exception as val_e:
                         print(f"    Warning: Also failed to get value for merge at ({min_row},{min_col}) on sheet '{sheet_name}'. Storing value as None. Error: {val_e}")
                         top_left_value = None
                     merges_data.append((col_span, top_left_value, None))

                except Exception as e:
                    print(f"    Warning: Could not get value/height for merge starting at ({min_row},{min_col}) on sheet '{sheet_name}'. Storing value/height as None. Error: {e}")
                    merges_data.append((col_span, None, None))

            original_merges[sheet_name] = merges_data
            print(f"  Stored {len(original_merges[sheet_name])} horizontal merge span/value/height entries for sheet '{sheet_name}'.")
            # Report skipped count for this filter
            if skipped_above_16_count > 0:
                print(f"    (Skipped {skipped_above_16_count} merges starting above row 16)")
        else:
             print(f"  Warning: Sheet '{sheet_name}' specified but not found during merge storage.")
             original_merges[sheet_name] = []
    return original_merges

# --- find_and_restore_merges_heuristic remains unchanged (still searches bottom-up, applies stored value/height) ---
def find_and_restore_merges_heuristic(workbook: openpyxl.Workbook,
                                      stored_merges: Dict[str, List[Tuple[int, Any, Optional[float]]]],
                                      processed_sheet_names: List[str],
                                      search_range_str: str = "A16:H200"):
    """
    Attempts to restore merges based on stored HORIZONTAL spans, values, and row heights
    by searching for the value within a specified range (default A16:H200).
    Searches rows bottom-up within the range.
    Includes detailed debugging output, prevents re-using a value, applies stored row height,
    and explicitly sets the stored value in the top-left cell after merging.

    WARNING: This is a HEURISTIC approach... (rest of docstring unchanged)

    Args: (args unchanged)
    """
    print(f"\nAttempting heuristic merge restoration (searching range {search_range_str}, bottom-up)...")
    restored_count = 0
    failed_count = 0
    skipped_count = 0
    skipped_duplicate_value_count = 0

    # --- Define search boundaries ---
    try:
        search_min_col, search_min_row, search_max_col, search_max_row = range_boundaries(search_range_str)
        print(f"  Search boundaries: Rows {search_min_row}-{search_max_row}, Cols {search_min_col}-{search_max_col} ({get_column_letter(search_min_col)}{search_min_row}:{get_column_letter(search_max_col)}{search_max_row})")
    except TypeError as te:
         print(f"  Error processing search range '{search_range_str}'. Check openpyxl version compatibility or range format. Internal error: {te}")
         traceback.print_exc()
         return
    except Exception as e:
        print(f"  Error: Invalid search range string '{search_range_str}'. Cannot proceed with restoration. Error: {e}")
        return
    # --- End boundary definition ---


    # --- Loop through sheets ---
    for sheet_name in processed_sheet_names:
        if sheet_name in workbook.sheetnames and sheet_name in stored_merges:
            worksheet: Worksheet = workbook[sheet_name]
            original_merges_data = stored_merges[sheet_name]
            print(f"  Processing sheet '{sheet_name}' ({len(original_merges_data)} stored merges)...")

            restored_start_cells = set()
            successfully_restored_values_on_sheet = set()

            # --- Loop through stored merge info ---
            for col_span, stored_value, stored_height in original_merges_data: # Unpack height

                if col_span <= 1:
                    skipped_count += 1
                    continue

                if stored_value in successfully_restored_values_on_sheet:
                    # print(f"    Skipping search for Value: '{stored_value}' (Span: {col_span}) - Value already used on sheet '{sheet_name}'.")
                    skipped_duplicate_value_count += 1
                    continue

                found = False
                # print(f"    Searching for Value: '{stored_value}' (Type: {type(stored_value)}), Target Span: {col_span}, Stored Height: {stored_height}")

                # --- Search range loop - ROW SEARCH REVERSED ---
                for r in range(search_max_row, search_min_row - 1, -1):
                    for c in range(search_min_col, search_max_col + 1):
                        cell_coord = (r, c)
                        if cell_coord in restored_start_cells:
                            continue

                        current_cell = worksheet.cell(row=r, column=c)
                        current_val = current_cell.value

                        # print(f"      Checking Cell {get_column_letter(c)}{r}: Value='{current_val}' (Type: {type(current_val)}) | Seeking: '{stored_value}' (Type: {type(stored_value)})")

                        if current_val == stored_value:
                            # print(f"        MATCH FOUND at {get_column_letter(c)}{r}!")
                            # print(f"    Attempting to merge {col_span} columns starting at {get_column_letter(c)}{r} for value '{stored_value}'.") # Keep this higher-level message
                            start_row, start_col = r, c
                            end_row = start_row
                            end_col = start_col + col_span - 1
                            target_range_str = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"

                            # --- Unmerge existing ---
                            merged_ranges_copy = list(worksheet.merged_cells.ranges)
                            for merged_range in merged_ranges_copy:
                                if merged_range.min_row <= start_row <= merged_range.max_row and \
                                   merged_range.min_col <= start_col <= merged_range.max_col:
                                     try:
                                         # print(f"      Unmerging existing range {merged_range.coord} overlapping target {target_range_str}")
                                         worksheet.unmerge_cells(str(merged_range))
                                     except KeyError: pass
                                     except Exception as ue: print(f"      Error unmerging existing range {merged_range.coord}: {ue}")

                            # --- Apply the new merge, Row Height, AND Value ---
                            try:
                                # 1. Apply merge
                                worksheet.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)
                                print(f"      Successfully merged {target_range_str}")

                                # 2. Apply stored row height
                                if stored_height is not None:
                                    try:
                                        worksheet.row_dimensions[start_row].height = stored_height
                                        print(f"      Applied row height {stored_height} to row {start_row}")
                                    except Exception as height_err:
                                        print(f"      Warning: Failed to apply row height {stored_height} to row {start_row}. Error: {height_err}")
                                else:
                                     print(f"      Stored height was None, row {start_row} keeps its current height.")

                                # 3. Restore the value to the top-left cell
                                try:
                                    top_left_cell_to_set = worksheet.cell(row=start_row, column=start_col)
                                    top_left_cell_to_set.value = stored_value
                                    print(f"      Set value '{stored_value}' to top-left cell {get_column_letter(start_col)}{start_row}")
                                except Exception as value_err:
                                    print(f"      Warning: Failed to set value '{stored_value}' to cell {get_column_letter(start_col)}{start_row}. Error: {value_err}")

                                # 4. Record success
                                restored_start_cells.add(cell_coord)
                                successfully_restored_values_on_sheet.add(stored_value)
                                restored_count += 1
                                found = True
                                break # Stop search loops for THIS stored_value pair

                            except Exception as e:
                                print(f"      Error merging cells, setting height, or setting value for {target_range_str}: {e}")
                                failed_count += 1
                                found = True # Still found, just failed
                                break # Stop search loops for THIS stored_value pair

                    if found:
                        break # Stop searching columns if found in current row

                # --- Check if found after loops ---
                if not found:
                    if stored_value not in successfully_restored_values_on_sheet:
                        # print(f"    -> Value '{stored_value}' (span {col_span}) NOT FOUND in range {search_range_str} on sheet '{sheet_name}'.")
                        failed_count += 1

        else:
            print(f"  Skipping merge restoration for sheet '{sheet_name}' (not found in workbook or no stored merges).")

    # --- Final Summary ---
    print("\nFinished heuristic merge restoration.")
    print(f"  Successfully restored: {restored_count}")
    print(f"  Failed/Not Found:    {failed_count}")
    print(f"  Skipped (span <= 1): {skipped_count}")
    print(f"  Skipped (value reused):{skipped_duplicate_value_count}")

# No __main__ block included as per previous request.
