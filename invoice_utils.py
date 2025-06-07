from pickle import NONE
import openpyxl
import re
import traceback
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill, NamedStyle
from openpyxl.utils import column_index_from_string, get_column_letter
from typing import List, Dict, Any, Optional, Tuple, Union
from decimal import Decimal

# --- Constants for Styling ---
thin_side = Side(border_style="thin", color="000000")
thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side) # Full grid border
no_border = Border(left=None, right=None, top=None, bottom=None)
center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
bold_font = Font(bold=True)

# --- Constants for Number Formats ---
FORMAT_GENERAL = 'General'
FORMAT_TEXT = '@'
FORMAT_NUMBER_COMMA_SEPARATED1 = '#,##0'
FORMAT_NUMBER_COMMA_SEPARATED2 = '#,##0.00'

# --- Utility Functions ---

def unmerge_row(worksheet: Worksheet, row_num: int, num_cols: int):
    """
    Unmerges any merged cells that overlap with the specified row within the given column range.

    Args:
        worksheet: The openpyxl Worksheet object.
        row_num: The 1-based row index to unmerge.
        num_cols: The number of columns to check for merges.
    """
    if row_num <= 0:
        return
    merged_ranges_copy = list(worksheet.merged_cells.ranges) # Copy ranges before modification
    merged_ranges_to_remove = []

    # Identify ranges that overlap with the target row
    for merged_range in merged_ranges_copy:
        # Check if the range's row span includes the target row_num
        # And if the range's column span overlaps with columns 1 to num_cols
        overlap = (merged_range.min_row <= row_num <= merged_range.max_row and
                   max(merged_range.min_col, 1) <= min(merged_range.max_col, num_cols))
        if overlap:
            merged_ranges_to_remove.append(str(merged_range))

    if merged_ranges_to_remove:
        for range_str in merged_ranges_to_remove:
            try:
                worksheet.unmerge_cells(range_str)
            except KeyError:
                # Range might have been removed by unmerging an overlapping one
                pass
            except Exception as unmerge_err:
                # Log or handle other potential errors if needed
                pass
    else:
        # No overlapping merges found for this row
        pass


def unmerge_block(worksheet: Worksheet, start_row: int, end_row: int, num_cols: int):
    """
    Unmerges any merged cells that overlap with the specified row range and column range.
    Args:
        worksheet: The openpyxl Worksheet object.
        start_row: The 1-based starting row index of the block.
        end_row: The 1-based ending row index of the block.
        num_cols: The number of columns to check for merges.
    """
    if start_row <= 0 or end_row < start_row:
        return
    merged_ranges_copy = list(worksheet.merged_cells.ranges) # Copy ranges before modification
    merged_ranges_to_remove = []

    # Identify ranges that overlap with the target block
    for merged_range in merged_ranges_copy:
        mr_min_row, mr_min_col, mr_max_row, mr_max_col = merged_range.bounds
        row_overlap = max(mr_min_row, start_row) <= min(mr_max_row, end_row)
        col_overlap = max(mr_min_col, 1) <= min(mr_max_col, num_cols)

        if row_overlap and col_overlap:
            range_str = str(merged_range)
            if range_str not in merged_ranges_to_remove: # Avoid duplicates
                merged_ranges_to_remove.append(range_str)

    if merged_ranges_to_remove:
        for range_str in merged_ranges_to_remove:
            try:
                worksheet.unmerge_cells(range_str)
            except KeyError:
                # Range might have been removed by unmerging an overlapping one
                pass
            except Exception as unmerge_err:
                # Log or handle other potential errors if needed
                pass
    else:
        # No overlapping merges found in this block
        pass


def safe_unmerge_block(worksheet: Worksheet, start_row: int, end_row: int, num_cols: int):
    """
    Safely unmerges only cells within the specific target range, preventing unintended unmerging
    of cells completely outside the block.
    """
    if start_row <= 0 or end_row < start_row:
        return

    # Only process merges that actually intersect with our target range
    for merged_range in list(worksheet.merged_cells.ranges):
        # Check if this merge intersects our target range
        if (merged_range.min_row <= end_row and
            merged_range.max_row >= start_row and
            merged_range.min_col <= num_cols and
            merged_range.max_col >= 1):
            try:
                worksheet.unmerge_cells(merged_range.coord)
            except (KeyError, ValueError, AttributeError):
                # Ignore errors if the range is somehow invalid or already unmerged
                continue

    return True


def fill_static_row(worksheet: Worksheet, row_num: int, num_cols: int, static_content_dict: Dict[str, Any]):
    """
    Fills a specific row with static content defined in a dictionary.
    Applies default alignment and no border.

    Args:
        worksheet: The openpyxl Worksheet object.
        row_num: The 1-based row index to fill.
        num_cols: The total number of columns in the table context (for bounds checking).
        static_content_dict: Dictionary where keys are column indices (as strings or ints)
                             and values are the static content to write.
    """
    if not static_content_dict:
        return # Nothing to do
    if row_num <= 0:
        return

    for col_key, value in static_content_dict.items():
        target_col_index = None
        try:
            # Attempt to convert key to integer column index
            target_col_index = int(col_key)
            # Check if the column index is within the valid range
            if 1 <= target_col_index <= num_cols:
                cell = worksheet.cell(row=row_num, column=target_col_index)
                cell.value = value
                # Apply default styling for static rows
                cell.alignment = center_alignment # Default alignment
                cell.border = no_border # Default: no border for static rows
                # Apply basic number formatting
                if isinstance(value, (int, float)):
                    cell.number_format = FORMAT_NUMBER_COMMA_SEPARATED2 if isinstance(value, float) else FORMAT_NUMBER_COMMA_SEPARATED1
                else:
                    cell.number_format = FORMAT_TEXT # Treat as text otherwise
            else:
                # Column index out of range, log warning?
                pass
        except (ValueError, TypeError) as e:
            # Invalid column key, log warning?
            pass
        except Exception as cell_err:
            # Error accessing cell, log warning?
            pass


def apply_row_merges(worksheet: Worksheet, row_num: int, num_cols: int, merge_rules: Optional[Dict[str, int]]):
    """
    Applies horizontal merges to a specific row based on rules.

    Args:
        worksheet: The openpyxl Worksheet object.
        row_num: The 1-based row index to apply merges to.
        num_cols: The total number of columns in the table context.
        merge_rules: Dictionary where keys are starting column indices (as strings or ints)
                     and values are the number of columns to span (colspan).
    """
    if not merge_rules or row_num <= 0:
        return # No rules or invalid row

    try:
        # Convert string keys to integers and sort for predictable application order
        rules_with_int_keys = {int(k): v for k, v in merge_rules.items()}
        sorted_keys = sorted(rules_with_int_keys.keys())
    except (ValueError, TypeError) as e:
        # Invalid key format in merge_rules
        return

    for start_col in sorted_keys:
        colspan_val = rules_with_int_keys[start_col]
        try:
            # Ensure colspan is an integer
            colspan = int(colspan_val)
        except (ValueError, TypeError):
            # Invalid colspan value
            continue

        # Basic validation for start column and colspan
        if not isinstance(start_col, int) or not isinstance(colspan, int) or start_col < 1 or colspan < 1:
            continue

        # Calculate end column, ensuring it doesn't exceed the table width
        end_col = start_col + colspan - 1
        if end_col > num_cols:
            end_col = num_cols
            # Check if clamping made the range invalid (start > end)
            if start_col > end_col:
                continue

        merge_range_str = f"{get_column_letter(start_col)}{row_num}:{get_column_letter(end_col)}{row_num}"
        try:
            # --- Pre-Unmerge Overlapping Cells ---
            merges_to_clear = []
            current_merged_ranges = list(worksheet.merged_cells.ranges) # Work on a copy
            for merged_range in current_merged_ranges:
                # Check if the existing merge overlaps with the target row and column range
                if merged_range.min_row <= row_num <= merged_range.max_row:
                    if max(merged_range.min_col, start_col) <= min(merged_range.max_col, end_col):
                        range_to_remove_str = str(merged_range)
                        if range_to_remove_str not in merges_to_clear:
                            merges_to_clear.append(range_to_remove_str)
            if merges_to_clear:
                for r_str in merges_to_clear:
                    try: worksheet.unmerge_cells(r_str)
                    except KeyError: pass
                    except Exception as unmerge_err_inner: pass # Log?
            # --- End Pre-Unmerge ---

            worksheet.merge_cells(start_row=row_num, start_column=start_col, end_row=row_num, end_column=end_col)
            # Apply alignment to the top-left cell of the merged range
            top_left_cell = worksheet.cell(row=row_num, column=start_col)
            if not top_left_cell.alignment or top_left_cell.alignment.horizontal is None:
                top_left_cell.alignment = center_alignment # Apply center alignment if none exists
        except ValueError as ve:
            # This can happen if trying to merge over an existing merged cell that wasn't properly unmerged
            pass
        except Exception as merge_err:
            # Log or handle other merge errors
            pass

def _apply_cell_style(cell, column_header: Optional[str], sheet_styling_config: Optional[Dict[str, Any]]):
    """
    Applies font, alignment, and number format to a cell based on default and column-specific rules.
    Respects explicitly set Text format.
    """
    if not sheet_styling_config or not cell:
        return # No styling config or invalid cell

    try:
        # --- Get Styling Configurations ---
        default_font_cfg = sheet_styling_config.get("default_font", {})
        default_align_cfg = sheet_styling_config.get("default_alignment", {})
        column_styles = sheet_styling_config.get("column_styles", {})

        # Find column-specific style rules if the header matches
        col_specific_style = {}
        if column_header and isinstance(column_styles, dict):
            col_specific_style = column_styles.get(column_header, {})

        # Extract font, alignment, and number format from column-specific rules (if they exist)
        col_font_cfg = col_specific_style.get("font", {}) if isinstance(col_specific_style, dict) else {}
        col_align_cfg = col_specific_style.get("alignment", {}) if isinstance(col_specific_style, dict) else {}
        number_format = col_specific_style.get("number_format") if isinstance(col_specific_style, dict) else None

        # --- Apply Font --- # RESTORED
        final_font_cfg = default_font_cfg.copy() if isinstance(default_font_cfg, dict) else {}
        if isinstance(col_font_cfg, dict): final_font_cfg.update(col_font_cfg) # Overwrite defaults
        if final_font_cfg:
            font_params = {k: v for k, v in final_font_cfg.items() if v is not None}
            if font_params:
                try: cell.font = Font(**font_params)
                except TypeError as e: pass # Ignore invalid font parameters
                except Exception as e_font: pass # Log other font errors?

        # --- Apply Alignment --- 
        # --- Apply Font --- # V12: IGNORED TO PRESERVE EXPLICITLY SET FONT (e.g., for footers)
        # final_font_cfg = default_font_cfg.copy() if isinstance(default_font_cfg, dict) else {}
        # if isinstance(col_font_cfg, dict): final_font_cfg.update(col_font_cfg) # Overwrite defaults
        # if final_font_cfg:
        #     font_params = {k: v for k, v in final_font_cfg.items() if v is not None}
        #     if font_params:
        #         try: cell.font = Font(**font_params)
        #         except TypeError as e: pass # Ignore invalid font parameters
        #         except Exception as e_font: pass # Log other font errors?

        # --- Apply Alignment --- 
        final_align_cfg = default_align_cfg.copy() if isinstance(default_align_cfg, dict) else {}
        if isinstance(col_align_cfg, dict): final_align_cfg.update(col_align_cfg) # Overwrite defaults
        if final_align_cfg:
            align_params = {k: v for k, v in final_align_cfg.items() if v is not None}
            if align_params:
                try: cell.alignment = Alignment(**align_params)
                except TypeError as e: pass # Ignore invalid alignment parameters
                except Exception as e_align: pass # Log other alignment errors?
        # --- Apply Number Format (Respecting Explicit Text Format) ---
        # Only apply number format if one is specified AND the cell format isn't already set to Text ('@')
        if number_format and cell.number_format != FORMAT_TEXT:
            try: cell.number_format = number_format
            except Exception as e_num_fmt: pass # Log number format errors?
        # Apply default number formats only if not Text and not already set by specific rule
        elif cell.number_format != FORMAT_TEXT and (cell.number_format == FORMAT_GENERAL or cell.number_format is None):
            if isinstance(cell.value, float): cell.number_format = FORMAT_NUMBER_COMMA_SEPARATED2
            elif isinstance(cell.value, int): cell.number_format = FORMAT_NUMBER_COMMA_SEPARATED1
            # No explicit else needed, keep General or existing format otherwise

    except Exception as style_err:
        # Log general styling errors?
        pass


def find_header(worksheet: Worksheet, header_rules: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    """
    Finds the header rows in a worksheet based on marker text and rules.
    Deprecated in favor of write_header for more control, but kept for potential legacy use.

    Args:
        worksheet: The openpyxl Worksheet object.
        header_rules: Dictionary defining how to find the header (markers, rows, logic).

    Returns:
        A dictionary containing header info (row indices, column map, num columns) or None if not found.
    """
    # --- DEPRECATED ---
    print("Warning: find_header is deprecated. Use write_header for better control.")
    markers = header_rules.get('markers', [])
    num_header_rows = header_rules.get('num_header_rows', 1)
    max_row_search = header_rules.get('max_row_to_search', worksheet.max_row)
    find_logic = header_rules.get('find_logic', 'or').lower() # 'or' or 'and'
    if not markers: return None
    if num_header_rows not in [1, 2]: return None

    first_row_index = -1; second_row_index = -1; num_columns = 0; column_map = {}
    try:
        # Find First Header Row
        for r_idx in range(1, min(max_row_search, worksheet.max_row) + 1):
            found_marker_in_row = False; markers_found_in_row = 0
            for marker_rule in markers:
                col_idx = marker_rule.get('column'); text_to_find = marker_rule.get('text')
                search_type = marker_rule.get('search_type', 'exact'); case_sensitive = marker_rule.get('case_sensitive', True)
                if not col_idx or text_to_find is None: continue
                try: cell = worksheet.cell(row=r_idx, column=col_idx); cell_value_str = str(cell.value) if cell.value is not None else ""
                except IndexError: continue
                found = False; text_to_find_str = str(text_to_find)
                if search_type == 'substring':
                    pattern = re.escape(text_to_find_str); flags = 0 if case_sensitive else re.IGNORECASE
                    if re.search(pattern, cell_value_str, flags): found = True
                elif case_sensitive and cell_value_str == text_to_find_str: found = True
                elif not case_sensitive and cell_value_str.lower() == text_to_find_str.lower(): found = True
                if found: found_marker_in_row = True; markers_found_in_row += 1;
                if find_logic == 'or': break
            if (find_logic == 'or' and found_marker_in_row) or (find_logic == 'and' and markers_found_in_row == len(markers)):
                first_row_index = r_idx; break
        if first_row_index == -1: return None

        # Determine Second Header Row
        if num_header_rows == 2:
            second_row_index = first_row_index + 1
            if second_row_index > worksheet.max_row: second_row_index = first_row_index # Fallback if at last row
            else: pass
        else: second_row_index = first_row_index

        # Determine Header Width (find last non-empty cell in the first header row)
        num_columns = 0
        for c_idx in range(worksheet.max_column, 0, -1):
            try:
                cell_val = worksheet.cell(row=first_row_index, column=c_idx).value
                if cell_val is not None and str(cell_val).strip() != "":
                    num_columns = c_idx
                    break
            except IndexError:
                continue # Skip if column index is out of bounds for this row
        if num_columns == 0: # Fallback if row is empty or error
            num_columns = worksheet.max_column
        pass

        # Build Column Map (Handle potential multi-row headers like "Quantity" -> "PCS", "SF")
        column_map = {}; header_row_1_vals = [worksheet.cell(row=first_row_index, column=c).value for c in range(1, num_columns + 1)];
        header_row_2_vals = []
        if num_header_rows == 2: header_row_2_vals = [worksheet.cell(row=second_row_index, column=c).value for c in range(1, num_columns + 1)]
        quantity_header_text = "Quantity"; sub_col_1_text = "PCS"; sub_col_2_text = "SF"; quantity_col_idx = -1
        if num_header_rows == 2:
            try: quantity_col_idx = header_row_1_vals.index(quantity_header_text) # Find 0-based index
            except (ValueError, TypeError): quantity_col_idx = -1
        for c_idx_0based in range(num_columns):
            col_index_1based = c_idx_0based + 1; header_text = None
            val_row1 = header_row_1_vals[c_idx_0based] if c_idx_0based < len(header_row_1_vals) else None
            val_row2 = header_row_2_vals[c_idx_0based] if c_idx_0based < len(header_row_2_vals) else None
            # Special handling for Quantity -> PCS/SF split
            if num_header_rows == 2 and quantity_col_idx != -1:
                if c_idx_0based == quantity_col_idx: # This is the PCS column
                    pcs_val = val_row2
                    if pcs_val is not None and str(pcs_val).strip() == sub_col_1_text: header_text = sub_col_1_text
                    else: header_text = quantity_header_text # Fallback if row 2 isn't PCS
                    # Check the *next* column in row 2 for SF
                    sf_col_index_0based = c_idx_0based + 1
                    if sf_col_index_0based < len(header_row_2_vals):
                        sf_val = header_row_2_vals[sf_col_index_0based]
                        if sf_val is not None and str(sf_val).strip() == sub_col_2_text:
                            column_map[sub_col_2_text] = col_index_1based + 1 # Map SF to the *next* column index
                    # Add the determined header (PCS or Quantity) to the map for the *current* column
                    if header_text: clean_header = str(header_text).strip();
                    if clean_header and clean_header not in column_map: column_map[clean_header] = col_index_1based
                    continue # Move to next column index
                elif c_idx_0based == quantity_col_idx + 1: # This is potentially the SF column
                    # If SF was already mapped in the previous step, skip this column
                    if sub_col_2_text in column_map and column_map[sub_col_2_text] == col_index_1based: continue
            # General case: Use row 2 if available, else row 1
            if num_header_rows >= 2 and val_row2 is not None and str(val_row2).strip() != "": header_text = str(val_row2).strip()
            elif val_row1 is not None and str(val_row1).strip() != "": header_text = str(val_row1).strip()
            # Add to map if a valid header text was found and not already mapped
            if header_text: clean_header = str(header_text).strip();
            if clean_header and clean_header not in column_map: column_map[clean_header] = col_index_1based
        if not column_map: return None # No valid headers found

        return {'first_row_index': first_row_index, 'second_row_index': second_row_index, 'column_map': column_map, 'num_columns': num_columns}
    except Exception as e: return None


def write_header(worksheet: Worksheet, start_row: int, header_data: List[List[Any]],
                 merge_rules: Optional[Dict[str, Any]] = None,
                 sheet_styling_config: Optional[Dict[str, Any]] = None
                 ) -> Optional[Dict[str, Any]]:
    if not header_data or not isinstance(header_data, list) or not all(isinstance(r, list) for r in header_data): return None
    num_header_rows = len(header_data)
    if num_header_rows == 0: return None
    num_columns = 0
    for r_data in header_data: num_columns = max(num_columns, len(r_data))
    if num_columns == 0: return None
    if start_row <= 0: return None
    end_row = start_row + num_header_rows - 1

    # --- Determine Header Styling ---
    # Default styles
    header_font_to_apply = bold_font       # Your defined default bold_font
    header_alignment_to_apply = center_alignment # Your defined default center_alignment
    header_border_to_apply = thin_border   # Your defined default thin_border
    header_background_fill_to_apply = None # Initialize to no specific background

    if sheet_styling_config:
        # Font configuration
        header_font_cfg = sheet_styling_config.get("header_font")
        if header_font_cfg and isinstance(header_font_cfg, dict):
            font_params = {k: v for k, v in header_font_cfg.items() if v is not None}
            if font_params:
                try: header_font_to_apply = Font(**font_params)
                except TypeError: pass

        # Alignment configuration
        header_align_cfg = sheet_styling_config.get("header_alignment")
        if header_align_cfg and isinstance(header_align_cfg, dict):
            align_params = {k: v for k, v in header_align_cfg.items() if v is not None}
            if align_params:
                try: header_alignment_to_apply = Alignment(**align_params)
                except TypeError: pass
        
        # --- Get and prepare header background fill ---
        header_fill_cfg = sheet_styling_config.get("header_pattern_fill") # Key from your JSON
        if header_fill_cfg and isinstance(header_fill_cfg, dict):
            fill_params = {k: v for k, v in header_fill_cfg.items() if v is not None}
            # Standardize keys for PatternFill constructor
            if "start_color" in fill_params and "fgColor" not in fill_params:
                fill_params["fgColor"] = fill_params.pop("start_color")
            if "end_color" in fill_params and "bgColor" not in fill_params:
                fill_params["bgColor"] = fill_params.pop("end_color")
            if "fill_type" in fill_params and "patternType" not in fill_params:
                fill_params["patternType"] = fill_params.pop("fill_type")
            
            if fill_params.get("fgColor") and fill_params.get("patternType"): # Essential params for PatternFill
                try:
                    header_background_fill_to_apply = PatternFill(**fill_params)
                except TypeError as e_hf_type:
                    print(f"Warning: Invalid header_pattern_fill parameters: {fill_params}. Error: {e_hf_type}")
                except Exception as e_hf:
                    print(f"Warning: Could not create header background fill: {e_hf}")
    try:
        unmerge_block(worksheet, start_row, end_row, num_columns)

        for r_offset, row_values in enumerate(header_data):
            current_row_idx = start_row + r_offset
            padded_row_values = row_values[:num_columns] + [None] * (num_columns - len(row_values))
            for c_idx_0based, value in enumerate(padded_row_values):
                col_index_1based = c_idx_0based + 1
                try:
                    cell = worksheet.cell(row=current_row_idx, column=col_index_1based)
                    cell.value = value
                    cell.font = header_font_to_apply
                    cell.alignment = header_alignment_to_apply
                    cell.border = header_border_to_apply
                    
                    # ***** THIS IS THE CORRECTION *****
                    if header_background_fill_to_apply:
                        cell.fill = header_background_fill_to_apply
                    # ***********************************
                        
                except Exception as write_err:
                    # print(f"Error writing header cell: {write_err}")
                    pass
        
        # --- Apply Vertical Merges ---
        if num_header_rows >= 2:
            for col_idx_1based in range(1, num_columns + 1):
                val1 = worksheet.cell(row=start_row, column=col_idx_1based).value
                if val1 is not None and str(val1).strip() != '':
                    should_merge_vertically = True
                    for r_offset in range(1, num_header_rows):
                        val_below = worksheet.cell(row=start_row + r_offset, column=col_idx_1based).value
                        if val_below is not None and str(val_below).strip() != '':
                            should_merge_vertically = False
                            break
                    if should_merge_vertically:
                        try:
                            merge_end_row = end_row # start_row + num_header_rows - 1
                            if start_row < merge_end_row: # Only merge if rowspan > 1
                                worksheet.merge_cells(start_row=start_row, start_column=col_idx_1based, end_row=merge_end_row, end_column=col_idx_1based)
                                # Ensure the anchor cell of the vertical merge has all styles
                                anchor_cell = worksheet.cell(row=start_row, column=col_idx_1based)
                                anchor_cell.alignment = header_alignment_to_apply
                                anchor_cell.font = header_font_to_apply # Ensure font is applied too
                                anchor_cell.border = header_border_to_apply # And border
                                if header_background_fill_to_apply: # And fill
                                     anchor_cell.fill = header_background_fill_to_apply
                        except Exception: # nosemgrep: general-exception-caught
                            pass 

        # --- Apply Horizontal/Colspan Merges (from merge_rules) ---
        if merge_rules:
            temp_col_map_written = {}
            first_header_row_idx = start_row
            for c in range(1, num_columns + 1):
                val = worksheet.cell(row=first_header_row_idx, column=c).value
                if val is not None: 
                    val_str = str(val).strip()
                    if val_str and val_str not in temp_col_map_written: 
                        temp_col_map_written[val_str] = c
            
            for header_text, merge_config in merge_rules.items():
                if isinstance(merge_config, dict):
                    start_col_idx = temp_col_map_written.get(header_text)
                    if start_col_idx:
                        colspan = int(merge_config.get('colspan', 1))
                        if colspan < 1: colspan = 1
                        rowspan = int(merge_config.get('rowspan', 1))
                        if rowspan < 1: rowspan = 1
                        if first_header_row_idx + rowspan - 1 > end_row: rowspan = end_row - first_header_row_idx + 1
                        
                        end_col_idx = start_col_idx + colspan - 1
                        end_row_idx = first_header_row_idx + rowspan - 1
                        if end_col_idx > num_columns: end_col_idx = num_columns

                        if end_col_idx >= start_col_idx and end_row_idx >= first_header_row_idx and (colspan > 1 or rowspan > 1):
                            try:
                                # Unmerge before merging is good practice, unmerge_block already called, 
                                # but specific smaller unmerges might be needed if rules overlap in complex ways.
                                # For now, assume initial unmerge_block is sufficient.
                                worksheet.merge_cells(start_row=first_header_row_idx, start_column=start_col_idx, end_row=end_row_idx, end_column=end_col_idx)
                                anchor_cell = worksheet.cell(row=first_header_row_idx, column=start_col_idx)
                                anchor_cell.alignment = header_alignment_to_apply
                                anchor_cell.font = header_font_to_apply # Ensure font
                                anchor_cell.border = header_border_to_apply # Ensure border
                                if header_background_fill_to_apply: # Ensure fill
                                     anchor_cell.fill = header_background_fill_to_apply
                            except ValueError: pass 
                            except Exception: pass

        # Rebuild Final Column Map (your existing logic)
        column_map_final = {}
        final_header_row_1_vals = [worksheet.cell(row=start_row, column=c).value for c in range(1, num_columns + 1)]
        final_header_row_2_vals = []
        final_second_row_index = start_row 
        if num_header_rows >= 2:
            final_second_row_index = start_row + 1
            final_header_row_2_vals = [worksheet.cell(row=final_second_row_index, column=c).value for c in range(1, num_columns + 1)]
        
        final_quantity_col_idx = -1
        if num_header_rows >= 2:
            try: final_quantity_col_idx = final_header_row_1_vals.index("Quantity")
            except (ValueError, TypeError): final_quantity_col_idx = -1
            
        for c_idx_0based in range(num_columns):
            col_index_1based = c_idx_0based + 1
            header_text_final = None
            val_row1 = final_header_row_1_vals[c_idx_0based] if c_idx_0based < len(final_header_row_1_vals) else None
            val_row2 = final_header_row_2_vals[c_idx_0based] if c_idx_0based < len(final_header_row_2_vals) else None

            if num_header_rows >= 2 and final_quantity_col_idx != -1:
                if c_idx_0based == final_quantity_col_idx: 
                    pcs_val = val_row2
                    if pcs_val is not None and str(pcs_val).strip() == "PCS": header_text_final = "PCS"
                    else: header_text_final = "Quantity"
                    sf_col_index_0based = c_idx_0based + 1
                    if sf_col_index_0based < len(final_header_row_2_vals):
                        sf_val = final_header_row_2_vals[sf_col_index_0based]
                        if sf_val is not None and str(sf_val).strip() == "SF":
                            column_map_final["SF"] = col_index_1based + 1
                    if header_text_final: 
                        clean_h = str(header_text_final).strip()
                        if clean_h and clean_h not in column_map_final: column_map_final[clean_h] = col_index_1based
                    continue 
                elif c_idx_0based == final_quantity_col_idx + 1: 
                    if "SF" in column_map_final and column_map_final["SF"] == col_index_1based: continue
            
            if num_header_rows >= 2 and val_row2 is not None and str(val_row2).strip() != "": header_text_final = str(val_row2).strip()
            elif val_row1 is not None and str(val_row1).strip() != "": header_text_final = str(val_row1).strip()
            
            if header_text_final: 
                clean_h_final = str(header_text_final).strip()
                if clean_h_final and clean_h_final not in column_map_final: column_map_final[clean_h_final] = col_index_1based
        
        if not column_map_final:
            # print("Warning: write_header resulted in an empty column_map_final.")
            # Fallback or decide how to handle. For now, let's try a basic map if num_columns > 0
            if num_columns > 0 and not header_data[0]: # If first row of header_data is empty, this won't work well
                 pass # Needs a better fallback if this case is possible
            elif num_columns > 0 and header_data[0]: # Try to use first row of header_data if column_map_final is empty
                for c_idx_0based, val_row1_fallback in enumerate(header_data[0][:num_columns]):
                    if val_row1_fallback is not None:
                        col_idx_1based_fb = c_idx_0based + 1
                        hdr_txt_fb = str(val_row1_fallback).strip()
                        if hdr_txt_fb and hdr_txt_fb not in column_map_final:
                            column_map_final[hdr_txt_fb] = col_idx_1based_fb
            if not column_map_final:
                 # print("Error: Still failed to build column_map_final in write_header.") # Already printed usually
                 return None


        return {'first_row_index': start_row, 
                'second_row_index': final_second_row_index, 
                'column_map': column_map_final, 
                'num_columns': num_columns}

    except Exception as e:
        # print(f"Error in write_header during main try block: {e}")
        # traceback.print_exc()
        return None


def find_footer(worksheet: Worksheet, footer_rules: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    """
    Finds the footer row based on marker text and rules.

    Args:
        worksheet: The openpyxl Worksheet object.
        footer_rules: Dictionary defining how to find the footer (marker text, columns, direction, etc.).

    Returns:
        A dictionary containing the footer start row {'start_row': index} or None if not found.
    """
    marker_text = footer_rules.get('marker_text'); search_type = footer_rules.get('search_type', 'exact'); case_sensitive = footer_rules.get('case_sensitive', True)
    search_columns = footer_rules.get('search_columns', [1]); search_direction = footer_rules.get('search_direction', 'down').lower()
    min_row_offset = footer_rules.get('min_row_offset', 1); max_row_search = footer_rules.get('max_row_to_search', worksheet.max_row)
    max_row_search = min(max_row_search, worksheet.max_row) # Ensure max_row_search doesn't exceed actual max row
    if not marker_text: return None
    if not isinstance(search_columns, list) or not search_columns: search_columns = [1]
    if min_row_offset <= 0: min_row_offset = 1

    try:
        # Determine Row Iteration Order
        row_iterator = None
        if search_direction == 'up': row_iterator = range(max_row_search, min_row_offset - 1, -1)
        else: row_iterator = range(min_row_offset, max_row_search + 1)
        marker_text_str = str(marker_text)

        # Search for Marker
        for r_idx in row_iterator:
            for c_idx in search_columns:
                if not (1 <= c_idx <= worksheet.max_column): continue # Skip invalid column index
                try:
                    cell = worksheet.cell(row=r_idx, column=c_idx)
                    # If it's a merged cell, only check the top-left origin cell of the merge range
                    if isinstance(cell, openpyxl.cell.cell.MergedCell):
                        is_origin = False
                        for merged_range in worksheet.merged_cells.ranges:
                            if merged_range.min_row == r_idx and merged_range.min_col == c_idx:
                                is_origin = True; break
                        if not is_origin: continue # Skip if not the top-left cell
                    cell_value_str = str(cell.value) if cell.value is not None else ""
                except IndexError: continue # Should not happen with max_column check, but safety first
                found = False
                if search_type == 'substring':
                    pattern = re.escape(marker_text_str); flags = 0 if case_sensitive else re.IGNORECASE
                    if re.search(pattern, cell_value_str, flags): found = True
                elif case_sensitive and cell_value_str == marker_text_str: found = True
                elif not case_sensitive and cell_value_str.lower() == marker_text_str.lower(): found = True
                if found: return {'start_row': r_idx} # Return immediately when found

        return None # Marker not found
    except Exception as e: return None # Error during search


# invoice_utils.py

# ... (imports, constants, other functions) ...

def write_configured_rows(
    worksheet: Worksheet,
    start_row_index: int,
    num_columns: int,
    rows_config_list: List[Dict[str, Any]], # Primary configuration for each row
    calculated_totals: Dict[str, Any],     # Data values to be inserted
    default_style_config: Optional[Dict[str, Any]] = None # Default styles from sheet config
):
    """
    Writes one or more rows with specified content (labels + dynamic values),
    styling, and merges based on configuration.
    Assumes the rows have already been inserted by the caller.
    Number formats are applied ONLY if specified in the cell's config,
    or set to Text ('@') for labels/non-numeric values.
    """
    if not rows_config_list or start_row_index <= 0:
        return

    print(f"--- Writing {len(rows_config_list)} configured rows starting at row {start_row_index} ---")
    calculated_totals = calculated_totals or {} # Ensure it's a dict

    # --- Get overall default styles from the sheet's styling configuration ---
    # These will be used if a row doesn't specify its own font/alignment.
    overall_default_font = Font() # Basic Openpyxl default
    overall_default_alignment = Alignment(horizontal='left', vertical='center', wrap_text=False) # Basic Openpyxl default

    if default_style_config:
        # Use 'default_font' and 'default_alignment' from the sheet's styling config if available
        sheet_default_font_cfg = default_style_config.get("default_font")
        if sheet_default_font_cfg and isinstance(sheet_default_font_cfg, dict):
            try:
                overall_default_font = Font(**{k: v for k, v in sheet_default_font_cfg.items() if v is not None})
            except TypeError:
                print("Warning: Invalid parameters in sheet's default_font config. Using basic default font.")
        
        sheet_default_align_cfg = default_style_config.get("default_alignment")
        if sheet_default_align_cfg and isinstance(sheet_default_align_cfg, dict):
            try:
                overall_default_alignment = Alignment(**{k: v for k, v in sheet_default_align_cfg.items() if v is not None})
            except TypeError:
                print("Warning: Invalid parameters in sheet's default_alignment config. Using basic default alignment.")

    # Iterate through each row's configuration object
    for i, row_config_item in enumerate(rows_config_list):
        current_row_idx = start_row_index + i
        print(f"  Processing configured row {i+1} (Sheet Row: {current_row_idx})")

        # --- Get ROW-LEVEL configurations from the current row_config_item ---
        row_cell_definitions = row_config_item.get("content", []) # List of cell configs for this row
        
        row_specific_height = row_config_item.get("height")
        row_specific_font_config = row_config_item.get("font")      # Font for the whole row
        row_specific_align_config = row_config_item.get("alignment") # Alignment for the whole row
        row_specific_merge_rules = row_config_item.get("merge_rules") # Merges for this specific row
        row_specific_apply_border = row_config_item.get("apply_default_border", True) # Border for the whole row

        # --- Determine effective font and alignment FOR THIS ENTIRE ROW ---
        # Start with the overall defaults, then apply row-level overrides if they exist.
        effective_row_font = overall_default_font
        if row_specific_font_config and isinstance(row_specific_font_config, dict):
            font_params = {k: v for k, v in row_specific_font_config.items() if v is not None}
            if font_params:
                try:
                    effective_row_font = Font(**font_params)
                except TypeError:
                    print(f"Warning: Invalid font config for row {current_row_idx}. Using sheet/basic default.")

        effective_row_alignment = overall_default_alignment
        if row_specific_align_config and isinstance(row_specific_align_config, dict):
            align_params = {k: v for k, v in row_specific_align_config.items() if v is not None}
            if align_params:
                try:
                    effective_row_alignment = Alignment(**align_params)
                except TypeError:
                    print(f"Warning: Invalid alignment config for row {current_row_idx}. Using sheet/basic default.")

        # --- Write Content Items (Cells) for the current row and Apply Styles ---
        written_columns_in_row = set() # Keep track of columns explicitly written to in this row
        
        if isinstance(row_cell_definitions, list):
            for cell_config_item in row_cell_definitions: # Each item in 'content' array from your JSON
                if not isinstance(cell_config_item, dict):
                    print(f"Warning: Invalid cell config item in row {current_row_idx}: {cell_config_item}")
                    continue

                try:
                    target_col_idx = int(cell_config_item.get("col"))
                    if not (1 <= target_col_idx <= num_columns):
                        print(f"Warning: Column index {target_col_idx} out of range for row {current_row_idx}.")
                        continue

                    cell = worksheet.cell(row=current_row_idx, column=target_col_idx)
                    written_columns_in_row.add(target_col_idx)
                    
                    value_to_write = None
                    # Cell-specific number format, font, and alignment from its own config
                    cell_specific_number_format = cell_config_item.get("number_format")
                    # Note: Cell-specific font/alignment could also be added to JSON if needed,
                    # otherwise, the effective_row_font/alignment will be used.

                    if "label" in cell_config_item:
                        value_to_write = cell_config_item["label"]
                        cell.number_format = cell_specific_number_format or FORMAT_TEXT # Use provided or default to Text
                    elif "value_key" in cell_config_item:
                        value_key = cell_config_item["value_key"]
                        raw_value = calculated_totals.get(value_key)
                        suffix = cell_config_item.get("suffix", "")
                        
                        numeric_value = None
                        if isinstance(raw_value, (int, float)):
                            numeric_value = float(raw_value)
                        elif isinstance(raw_value, str):
                            try:
                                # Attempt to convert if it looks like a number, handling commas
                                cleaned_raw_value = raw_value.replace(',', '')
                                if cleaned_raw_value.strip(): # Avoid empty strings
                                    numeric_value = float(cleaned_raw_value)
                            except (ValueError, TypeError):
                                pass # Keep as None if conversion fails

                        if numeric_value is not None:
                            # If there's a suffix, the value becomes a string.
                            # If no suffix, keep it as a number for Excel to handle.
                            value_to_write = f"{numeric_value}{suffix}" if suffix else numeric_value
                            
                            if cell_specific_number_format:
                                cell.number_format = cell_specific_number_format
                            elif suffix: # If suffix is present, it's text
                                cell.number_format = FORMAT_TEXT
                            # Else (numeric, no suffix, no specific format): Let Excel use default number format
                            
                        else: # Value is not numeric or not found, treat as text
                            value_to_write = f"{str(raw_value or '')}{suffix}" # Use empty string if raw_value is None
                            cell.number_format = cell_specific_number_format or FORMAT_TEXT
                    
                    elif "value" in cell_config_item: # Direct static value
                        value_to_write = cell_config_item.get("value")
                        # Assume direct static values are text unless a number_format is given
                        cell.number_format = cell_specific_number_format or FORMAT_TEXT
                    else:
                        # No label, value_key, or value; cell might be intended to be blank but styled
                        pass


                    cell.value = value_to_write
                    cell.font = effective_row_font # Apply the determined row font
                    cell.alignment = effective_row_alignment # Apply the determined row alignment

                    # Apply border based on row-level setting
                    if row_specific_apply_border:
                        cell.border = thin_border
                    else:
                        cell.border = no_border

                except (ValueError, TypeError) as e:
                    print(f"Warning: Invalid data in cell config for row {current_row_idx}: {cell_config_item}. Error: {e}")
                except Exception as cell_err:
                    print(f"Warning: Error writing cell (Row: {current_row_idx}, Col: {cell_config_item.get('col', 'N/A')}): {cell_err}")

        # --- Ensure remaining (unwritten) cells in the row get default row styling (border) ---
        for c_idx_fill in range(1, num_columns + 1):
            if c_idx_fill not in written_columns_in_row: # Only touch columns not explicitly defined
                try:
                    cell = worksheet.cell(row=current_row_idx, column=c_idx_fill)
                    # Apply row's effective font and alignment to blank cells if desired (optional)
                    # cell.font = effective_row_font
                    # cell.alignment = effective_row_alignment
                    if row_specific_apply_border:
                        cell.border = thin_border
                    else:
                        # Only remove border if cell is truly blank and no border is intended for the row
                        if cell.value is None: # Check if cell is actually empty
                            cell.border = no_border
                except Exception as blank_cell_err:
                    print(f"Warning: Error styling blank cell ({current_row_idx},{c_idx_fill}): {blank_cell_err}")


        # --- Apply Merges for this entire row (using row-level merge rules) ---
        if row_specific_merge_rules and isinstance(row_specific_merge_rules, dict):
            apply_row_merges(worksheet, current_row_idx, num_columns, row_specific_merge_rules)
            # Re-apply style/border to the top-left cell of any merged ranges
            # to ensure consistent appearance, as merging can sometimes affect the primary cell's style.
            for start_col_str_merge in row_specific_merge_rules.keys():
                try:
                    start_col_idx_merge = int(start_col_str_merge)
                    merged_cell_anchor = worksheet.cell(row=current_row_idx, column=start_col_idx_merge)
                    merged_cell_anchor.font = effective_row_font
                    merged_cell_anchor.alignment = effective_row_alignment
                    if row_specific_apply_border:
                        merged_cell_anchor.border = thin_border
                    else:
                        merged_cell_anchor.border = no_border
                except (ValueError, TypeError):
                    print(f"Warning: Invalid start column for merge rule on row {current_row_idx}: {start_col_str_merge}")
                except Exception as merge_style_err:
                    print(f"Warning: Error re-styling merged cell anchor at ({current_row_idx},{start_col_str_merge}): {merge_style_err}")

        # --- Apply Height for this entire row (using row-level height) ---
        if row_specific_height is not None:
            try:
                h_val = float(row_specific_height)
                if h_val > 0:
                    worksheet.row_dimensions[current_row_idx].height = h_val
            except (ValueError, TypeError):
                print(f"Warning: Invalid height value '{row_specific_height}' for row {current_row_idx}.")
            except Exception as height_err:
                print(f"Warning: Error setting height for row {current_row_idx}: {height_err}")

    print(f"--- Finished writing configured rows ---")

def apply_explicit_data_cell_merges(
    worksheet: Worksheet,
    row_num: int,
    column_map: Dict[str, int],  # Maps header text to its 1-based column index
    num_total_columns: int,
    # merge_rules_data_cells: e.g., {'Header': {'rowspan': 3, ...}}
    merge_rules_data_cells: Dict[str, Dict[str, Any]], 
    sheet_styling_config: Optional[Dict[str, Any]] # e.g. {'Header': {'font': {'bold':True}}}
):
    """
    Applies horizontal merges to data cells in a specific row
    based on explicit rules. The merged cell will have a thin black border
    around its entire perimeter and its content will be centered.
    """
    if not merge_rules_data_cells or row_num <= 0:
        return

    thin_side = Side(border_style="thin", color="000000")
    full_thin_border = Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side
    )
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for header_text, rule_details in merge_rules_data_cells.items():
        colspan_to_apply = rule_details.get("rowspan") 

        if not isinstance(colspan_to_apply, int) or colspan_to_apply <= 1:
            continue
        start_col_idx = column_map.get(header_text)
        if not start_col_idx:
            continue
        end_col_idx = start_col_idx + colspan_to_apply - 1
        end_col_idx = min(end_col_idx, num_total_columns)

        if start_col_idx >= end_col_idx:
            continue

        try:
            current_merged_cell_ranges = list(worksheet.merged_cells.ranges)
            for mc_range in current_merged_cell_ranges:
                if mc_range.min_row == row_num and mc_range.max_row == row_num:
                    if mc_range.min_col <= end_col_idx and mc_range.max_col >= start_col_idx:
                        try:
                            worksheet.unmerge_cells(str(mc_range))
                        except Exception:
                            pass
            
            worksheet.merge_cells(start_row=row_num, start_column=start_col_idx,
                                  end_row=row_num, end_column=end_col_idx)
            
            anchor_cell: Cell = worksheet.cell(row=row_num, column=start_col_idx)
            
            # === DEBUGGING STEP FOR BORDERS ===
            # If borders are not appearing AT ALL, try commenting out the following
            # call to _apply_cell_style to isolate the issue. If borders appear
            # when this is commented out, then _apply_cell_style is likely resetting
            # or interfering with the border style.
            if sheet_styling_config: 
                 specific_rules_for_cell = sheet_styling_config.get(header_text)
                 # _apply_cell_style(anchor_cell, header_text, specific_rules_for_cell) # <-- TRY COMMENTING THIS LINE
            # === END DEBUGGING STEP ===

            anchor_cell.border = full_thin_border
            anchor_cell.alignment = center_alignment

        except Exception as e:
            print(f"Error applying explicit data cell merge, border, or alignment for '{header_text}' on row {row_num} (col {start_col_idx}, span {colspan_to_apply}): {e}")




def fill_invoice_data(
    worksheet: Worksheet,
    sheet_name: str,
    sheet_config: Dict[str, Any], # Keep current sheet config param
    all_sheet_configs: Dict[str, Any], # <--- Add param for all sheet configs
    data_source: Union[Dict[str, List[Any]], Dict[Tuple, Dict[str, Any]]],
    data_source_type: str,
    header_info: Dict[str, Any],
    mapping_rules: Dict[str, Any],
    sheet_styling_config: Optional[Dict[str, Any]] = None,
    add_blank_after_header: bool = False,
    static_content_after_header: Optional[Dict[str, Any]] = None,
    add_blank_before_footer: bool = False,
    static_content_before_footer: Optional[Dict[str, Any]] = None,
    merge_rules_after_header: Optional[Dict[str, int]] = None,
    merge_rules_before_footer: Optional[Dict[str, int]] = None,
    merge_rules_footer: Optional[Dict[str, int]] = None, # Added footer merge rules
    footer_info: Optional[Dict[str, Any]] = None, # Currently unused
    max_rows_to_fill: Optional[int] = None,
    grand_total_pallets: int = 0, # RE-ADDED parameter
    custom_flag: bool = False, # Added custom flag parameter
    data_cell_merging_rules: Optional[Dict[str, Any]] = None # Added data cell merging rules 29/05/2025
    ) -> Tuple[bool, int, int, int, int]: # Still 5 return values
    """
    REVISED LOGIC V13: Added merge_rules_footer parameter.
    Footer pallet count uses local_chunk_pallets for processed_tables,
    and grand_total_pallets for aggregation/fob_aggregation.
    """

    # --- Initialize Variables --- (Keep existing initializations)
    actual_rows_to_process = 0; data_rows_prepared = []; col1_index = 1; num_static_labels = 0
    static_column_header_name = None; data_row_indices_written = [];
    columns_to_grid = []
    desc_col_idx = None
    local_chunk_pallets = 0
    dynamic_desc_used = False



    # get data source pallet count and hanle null
    for pallet_count in data_source.get("pallet_count", []):
        if pallet_count is not None:
            local_chunk_pallets += pallet_count

    # --- Row Index Tracking --- (Keep existing)
    row_after_header_idx = -1
    data_start_row = -1
    data_end_row = -1
    row_before_footer_idx = -1
    footer_row_final = -1

    # Ensure dictionaries/lists are initialized (Keep existing)
    static_content_after_header = static_content_after_header or {}
    static_content_before_footer = static_content_before_footer or {}
    merge_rules_after_header = merge_rules_after_header or {}
    merge_rules_before_footer = merge_rules_before_footer or {}
    merge_rules_footer = merge_rules_footer or {} # Initialize footer merge rules
    mapping_rules = mapping_rules or {}

    try:
        data_cell_merging_rules = data_cell_merging_rules or {}
        # --- Validate Header Info --- (Keep existing)
        if not header_info or 'second_row_index' not in header_info or 'column_map' not in header_info or 'num_columns' not in header_info:
            print("Error: Invalid header_info provided.")
            return False, -1, -1, -1, 0
        # V11: Determine start point based *directly* on passed header info
        # initial_insert_point = header_info['second_row_index'] + 1 # OLD LOGIC
        data_writing_start_row = header_info['second_row_index'] + 1 # Where data *content* begins
 
        column_map = header_info['column_map']; num_columns = header_info['num_columns']
        idx_to_header_map = {v: k for k, v in column_map.items()}

        # --- Find Description & Pallet Info Column Indices --- (Keep existing)
        possible_desc_headers = ["Description\ntả hàng hóa","Description", "Description of Goods", "DESCRIPTION OF GOODS", "DESCRIPTION"]
        desc_col_idx = None
        for h in possible_desc_headers:
            for map_key, map_idx in column_map.items():
                 if isinstance(map_key, str) and map_key.strip().lower() == h.lower(): desc_col_idx = map_idx; break
            if desc_col_idx is not None: break
        if desc_col_idx is None: print("Warning: Could not find 'Description' column header.")

        pallet_info_col_idx = column_map.get("Pallet\nNo")
        if pallet_info_col_idx is None: print("Warning: Header 'Pallet Info' not found.")

        # --- ADD/MODIFY THIS PART FOR PALLET INFO INDEX ---
        possible_pallet_headers = ["Pallet\nNo", "Pallet No", "PALLET\nNO", "PALLETINFO", "Pallet Information", "PALLET\nNO."] # Add other variations as needed
        pallet_info_col_idx = None
        for h_pallet in possible_pallet_headers:
            for map_key_pallet, map_idx_pallet in column_map.items():
                if isinstance(map_key_pallet, str) and map_key_pallet.strip().lower() == h_pallet.lower():
                    pallet_info_col_idx = map_idx_pallet
                    break
            if pallet_info_col_idx is not None:
                break
        if pallet_info_col_idx is None:
            print("Warning: Could not find a 'Pallet Info' (e.g., 'Pallet\\nNo') column header.")
        # --- END OF ADDITION/MODIFICATION FOR PALLET INFO INDEX ---

        # --- Get Styling Config --- (Keep existing)
        force_text_headers = []
        effective_header_font = bold_font # Start with default
        effective_header_align = center_alignment # Start with default

        if sheet_styling_config:
            columns_to_grid = sheet_styling_config.get("columns_with_full_grid", [])
            if not isinstance(columns_to_grid, list): columns_to_grid = []

            force_text_headers = sheet_styling_config.get("force_text_format_headers", [])
            if not isinstance(force_text_headers, list): force_text_headers = []

            header_font_cfg = sheet_styling_config.get("header_font")
            if header_font_cfg and isinstance(header_font_cfg, dict):
                 font_params = {k: v for k, v in header_font_cfg.items() if v is not None}
                 if font_params:
                     try: # Expanded try block
                         effective_header_font = Font(**font_params)
                     except TypeError:
                         print(f"Warning: Invalid parameters in header_font config: {font_params}. Using default.")
                         pass # Keep default font on error
                     except Exception as font_err: # Catch other potential errors
                         print(f"Warning: Error applying header_font config: {font_err}. Using default.")
                         pass # Keep default font on error

            header_align_cfg = sheet_styling_config.get("header_alignment")
            if header_align_cfg and isinstance(header_align_cfg, dict):
                 align_params = {k: v for k, v in header_align_cfg.items() if v is not None}
                 if align_params:
                     try: # Expanded try block
                         effective_header_align = Alignment(**align_params)
                     except TypeError:
                         print(f"Warning: Invalid parameters in header_alignment config: {align_params}. Using default.")
                         pass # Keep default alignment on error
                     except Exception as align_err: # Catch other potential errors
                          print(f"Warning: Error applying header_alignment config: {align_err}. Using default.")
                          pass # Keep default alignment on error

        # --- Prepare Data Mapping Rules (Separate Static, Initial Static, Dynamic, and **FORMULA**) ---
        static_value_map = {}; initial_static_col1_values = []; dynamic_mapping_rules = {}; initial_static_rule = None
        # This dictionary will store rules specifically marked as "type": "formula" in the config.
        # Keys will be the target column index, values will contain the template and input headers.
        formula_rules = {}

        # First pass: find the initial static rule (Keep existing)
        for rule_key, rule_value in mapping_rules.items():
            if isinstance(rule_value, dict) and rule_value.get("type") == "initial_static_rows":
                 initial_static_rule = rule_value; static_column_header_name = initial_static_rule.get("column_header")
                 target_col_idx = column_map.get(static_column_header_name) if static_column_header_name else None
                 if target_col_idx: col1_index = target_col_idx; initial_static_col1_values = initial_static_rule.get("values", []); num_static_labels = len(initial_static_col1_values)
                 else: print(f"Warning: Initial static rows column '{static_column_header_name}' not found.")
                 break

        # Second pass: separate static, dynamic, and formula rules
        for rule_key, rule_value in mapping_rules.items():
            if isinstance(rule_value, dict):
                rule_type = rule_value.get("type")
                header_text = rule_value.get("header")
                target_col_idx = column_map.get(header_text) if header_text else None

                if rule_type == "initial_static_rows": continue # Skip the one found above
                if 'marker' in rule_value: continue # Skip summary field markers

                # --- Formula Rule Processing ---
                # If a mapping rule in the config has "type": "formula", store its details
                # (template string and list of input header names) keyed by the target column index.
                if rule_type == "formula":
                    if target_col_idx:
                        formula_template = rule_value.get("formula_template")
                        input_headers = rule_value.get("inputs")
                        if formula_template and isinstance(input_headers, list):
                            formula_rules[target_col_idx] = {
                                "template": formula_template,
                                "input_headers": input_headers
                            }
                            print(f"DEBUG: Parsed formula rule for header '{header_text}' (Col {target_col_idx})")
                        else:
                            print(f"Warning: Invalid formula rule for header '{header_text}'. Missing template or inputs.")
                    else:
                        print(f"Warning: Could not find target column for formula rule with header '{header_text}'.")
                # --- End Formula Rule Processing ---

                elif "static_value" in rule_value: # Existing static value logic
                    if target_col_idx: static_value_map[target_col_idx] = rule_value["static_value"]
                else: # Existing dynamic rule logic (dict without static_value and not formula)
                    dynamic_mapping_rules[rule_key] = rule_value
            else: # Existing simple string mapping (dynamic)
                dynamic_mapping_rules[rule_key] = rule_value

        apply_special_border_rule = static_column_header_name and static_column_header_name.strip() in ["Mark & Nº", "Mark & N °"]

        # --- Prepare Data Rows for Writing (Determine number of rows needed from source) ---
        # This section remains largely the same, preparing the `data_rows_prepared` list
        # which holds the *input* data, not the calculated formulas.
        num_data_rows_from_source = 0
        pallet_counts_for_rows = []

        if data_source_type == 'processed_tables':
            # ... (Keep existing logic for processed_tables data preparation) ...
            # Ensure 'pallet_count' logic is still correct here.
            # The special Description fallback logic also remains here.
             data_source = data_source or {}; max_len = 0
             if isinstance(data_source, dict):
                 for value in data_source.values():
                     if isinstance(value, list): max_len = max(max_len, len(value))
                 num_data_rows_from_source = max_len
                 raw_pallet_counts = data_source.get("pallet_count", [])
                 if isinstance(raw_pallet_counts, list): pallet_counts_for_rows = raw_pallet_counts[:max_len] + [0] * (max_len - len(raw_pallet_counts))
                 else: pallet_counts_for_rows = [0] * max_len; print("Warning: 'pallet_count' key missing...")
                 if num_data_rows_from_source > 0:
                     for i in range(num_data_rows_from_source):
                         row_dict = {}
                         # Iterate through dynamic rules (which map json keys to headers/rules)
                         for json_key, header_or_rule in dynamic_mapping_rules.items():
                             target_col_idx = None
                             data_value_from_source = None # Store original value
                             effective_value_to_write = None # Final value after fallback
                             is_description_column = False
                             mapping_rule = None # Store the specific rule dict if applicable

                             # --- Determine Target Column and Mapping Rule --- #
                             if isinstance(header_or_rule, str):
                                 # Simple case: Rule is just the header text
                                 header_text = header_or_rule
                                 target_col_idx = column_map.get(header_text)
                                 # Need to find the corresponding rule dict for fallback
                                 for r_key, r_val in mapping_rules.items():
                                     if isinstance(r_val, dict) and r_val.get("header") == header_text:
                                         mapping_rule = r_val
                                         break
                             elif isinstance(header_or_rule, dict) and "static_value" not in header_or_rule and header_or_rule.get("type") != "formula":
                                 # Complex case: Rule is a dictionary
                                 header_text = header_or_rule.get("header")
                                 target_col_idx = column_map.get(header_text) if header_text else None
                                 mapping_rule = header_or_rule # The rule is directly available

                             # --- Get Value from Source Data --- #
                             if target_col_idx:
                                 source_list = data_source.get(json_key)
                                 if isinstance(source_list, list) and i < len(source_list):
                                     data_value_from_source = source_list[i]
                                 is_description_column = (target_col_idx == desc_col_idx)

                                 # --- Determine Effective Value (Apply Fallback) --- #
                                 effective_value_to_write = None # Initialize
                                 value_was_determined = False # Flag to track if any value (even None) was explicitly determined

                                 # Check if source value is considered "empty" (None or whitespace string)
                                 is_empty_source_value = (data_value_from_source is None or 
                                                         (isinstance(data_value_from_source, str) and not data_value_from_source.strip()))

                                 if not is_empty_source_value:
                                     effective_value_to_write = data_value_from_source
                                     value_was_determined = True
                                     if is_description_column: dynamic_desc_used = True # Mark if actual data used for description
                                 else:
                                     # Source value is empty, try fallback
                                     fallback_value = None
                                     if mapping_rule and isinstance(mapping_rule, dict):
                                         # Use .get() which returns None if key doesn't exist
                                         fallback_value = mapping_rule.get("fallback_on_none")
                                     
                                     if fallback_value is not None: # Check if fallback *exists*
                                         effective_value_to_write = fallback_value
                                         value_was_determined = True
                                         # If description uses fallback, don't mark dynamic_desc_used
                                     elif is_description_column:
                                         # Last resort for description: check static_value_map
                                         static_map_value = static_value_map.get(target_col_idx)
                                         if static_map_value is not None: # Check if static value exists
                                             effective_value_to_write = static_map_value
                                             value_was_determined = True
                                     # else: effective_value_to_write remains None for other columns if no fallback
                                     # and value_was_determined remains False

                                 # --- Add the key to the row dictionary if a value was determined --- #
                                 # This ensures the key exists even if the determined value is None or ""
                                 if value_was_determined:
                                     row_dict[target_col_idx] = effective_value_to_write
                             # else: target_col_idx not found for this rule, skip

                         # --- Ensure Description column exists even if all sources/fallbacks were None --- #
                         # This check might be less necessary now but acts as a final safety net
                         if desc_col_idx is not None and desc_col_idx not in row_dict:
                             # Check mapping rules again for the description column specifically
                             final_fallback_desc = None
                             desc_header = idx_to_header_map.get(desc_col_idx)
                             if desc_header:
                                 for rule in mapping_rules.values():
                                     if isinstance(rule, dict) and rule.get('header') == desc_header:
                                         final_fallback_desc = rule.get('fallback_on_none')
                                         if final_fallback_desc is None:
                                             final_fallback_desc = rule.get('static_value')
                                         break # Found rule
                             # If still nothing, check static map (original fallback)
                             if final_fallback_desc is None:
                                 final_fallback_desc = static_value_map.get(desc_col_idx)

                             # Add to dict ONLY if a final fallback was actually found
                             if final_fallback_desc is not None:
                                 row_dict[desc_col_idx] = final_fallback_desc

                         # Append the prepared row data
                         if row_dict or i < num_static_labels: # Keep row if it has data OR it's for an initial static label
                            data_rows_prepared.append(row_dict)
                         elif i >= num_static_labels: # Only append empty dict if past static labels and row is truly empty
                            data_rows_prepared.append({}) 

        elif data_source_type == 'aggregation':
            # ... (Keep existing logic for aggregation data preparation) ...
            # Ensure fallback logic is correct.
            # Pallet count assumption remains.
             print("Warning: Pallet count per row is assumed '1' for 'aggregation' data source type.")
             num_data_rows_from_source = len(data_source) if isinstance(data_source, dict) else 0
             pallet_counts_for_rows = [1] * num_data_rows_from_source # Assume 1 pallet
             if num_data_rows_from_source > 0 and isinstance(data_source, dict):
                 row_counter = 0
                 for key_tuple, value_dict in data_source.items():
                     row_dict = {}
                     for map_key, map_rule in dynamic_mapping_rules.items():
                         if not isinstance(map_rule, dict) or 'static_value' in map_rule or 'marker' in map_rule or map_rule.get("type") in ["initial_static_rows", "formula"]: continue # Exclude formula rules here
                         header_text = map_rule.get('header'); target_col_index = column_map.get(header_text) if header_text else None;
                         if not target_col_index: continue
                         data_value = None; original_value_is_none = True
                         try:
                             if 'key_index' in map_rule:
                                 key_index = map_rule['key_index']
                                 if isinstance(key_tuple, tuple) and isinstance(key_index, int) and 0 <= key_index < len(key_tuple):
                                     data_value = key_tuple[key_index]; original_value_is_none = (data_value is None)
                                     if original_value_is_none and "fallback_on_none" in map_rule: data_value = map_rule["fallback_on_none"]
                                 elif "fallback_on_none" in map_rule: data_value = map_rule["fallback_on_none"]
                             elif 'value_key' in map_rule:
                                 value_key = map_rule['value_key']
                                 if isinstance(value_dict, dict) and value_key in value_dict:
                                     data_value = value_dict[value_key]; original_value_is_none = (data_value is None)
                         except Exception as e: pass
                         if data_value is not None:
                             row_dict[target_col_index] = data_value
                             if target_col_index == desc_col_idx and not original_value_is_none: dynamic_desc_used = True
                     for static_col_idx, static_val in static_value_map.items():
                         if static_col_idx not in row_dict: row_dict[static_col_idx] = static_val
                     if row_dict or row_counter < num_static_labels: data_rows_prepared.append(row_dict)
                     else: data_rows_prepared.append({})
                     row_counter += 1

                     # --- Custom Aggregation Amount Override (Before Appending) ---
                     # This block overrides the standard mapping for Amount if custom_flag is true
                     if custom_flag:
                         print(f"DEBUG: Custom flag True. Overriding Amount for aggregation row {row_counter}.")
                         # Find Amount column index
                         amount_col_idx_override = None
                         amount_headers_override = ["Amount ( USD )", "Total value(USD)", "amount", "amount_sum", "Amount(USD)"]
                         for header, col_idx in column_map.items():
                             if str(header).lower() in [h.lower() for h in amount_headers_override]:
                                 amount_col_idx_override = col_idx
                                 break

                         # Get amount_sum from the value dictionary
                         amount_sum_value = None
                         if isinstance(value_dict, dict):
                             amount_sum_value = value_dict.get("amount_sum")

                         if amount_col_idx_override and amount_sum_value is not None:
                             print(f"DEBUG: Custom Override - Found Amount Col: {amount_col_idx_override}, Amount Sum: {amount_sum_value}")
                             # Convert to float for consistency before writing
                             try:
                                 amount_float = float(str(amount_sum_value).replace(',', ''))
                                 row_dict[amount_col_idx_override] = amount_float # Overwrite the mapped value
                                 print(f"DEBUG: Custom Override - Updated row_dict[{amount_col_idx_override}] = {amount_float}")
                             except (ValueError, TypeError) as e:
                                 print(f"Warning: Custom Override - Could not convert amount_sum '{amount_sum_value}' to float: {e}. Keeping original mapped value if any.")
                         else:
                             missing_info = []
                             if not amount_col_idx_override: missing_info.append("Amount column in header map")
                             if amount_sum_value is None: missing_info.append('"amount_sum" key in data value')
                             print(f"Warning: Custom Override - Skipping amount override. Missing: {', '.join(missing_info)}")
                     # --- End Custom Aggregation Amount Override ---

        elif data_source_type == 'fob_aggregation':
            # --- FOB Aggregation Data Prep (Revised for Nested Dict) ---
            # Prepare data_rows_prepared based on the new structure.
            # Also find initial static labels.
            num_static_labels = 0
            initial_static_col1_values = []
            col1_index = 1 # Default
            static_column_header_name = None
            data_rows_prepared = [] # Initialize here for FOB

            # Find initial static rule (same as before)
            for rule_key, rule_value in mapping_rules.items():
                if isinstance(rule_value, dict) and rule_value.get("type") == "initial_static_rows":
                    static_column_header_name = rule_value.get("column_header")
                    target_col_idx = column_map.get(static_column_header_name) if static_column_header_name else None
                    if target_col_idx:
                        col1_index = target_col_idx
                        initial_static_col1_values = rule_value.get("values", [])
                        num_static_labels = len(initial_static_col1_values)
                    break # Assume only one such rule

            print(f"DEBUG: FOB Mode - Found {num_static_labels} initial static labels configured.")

            # Prepare data rows from the nested data_source dictionary
            if data_source and isinstance(data_source, dict):
                fob_data_keys = { # Mapping from Sheet Header -> Data Source Key
                    'P.O N °': 'combined_po', 'P.O Nº': 'combined_po', 'P.ON°': 'combined_po',
                    'ITEM NO': 'combined_item', 'ITEM Nº': 'combined_item' , 'ITEM N°': 'combined_item', 'Name of Cormodity': 'combined_item',
                    "Name of\nCormodity": "combined_item",
                    'Quantity ( SF )': 'total_sqft', 'Quantity(SF)': 'total_sqft', "Quantity(SF)": 'total_sqft', "Unit Price(USD)": 'unit_price',
                    'Amount ( USD )': 'total_amount', 'Total value(USD)': 'total_amount', "P.O Nº": "combined_po", "P.O N°": "combined_po",
                    "Quantity\n(SF)": 'total_sqft', "Amount(USD)": 'total_amount', "Description Of Goods": "combined_item",
                }
                desc_header_options = ["Description", "DESCRIPTION OF GOODS", "Description of Goods", "DESCRIPTION"]
                desc_col_idx_fob_prep = None
                for header in desc_header_options:
                    desc_col_idx_fob_prep = column_map.get(header)
                    if desc_col_idx_fob_prep: break

                # Sort by key to ensure order ("1", "2", ...)
                sorted_keys = sorted(data_source.keys(), key=lambda k: int(k) if k.isdigit() else float('inf'))

                for row_key in sorted_keys:
                    row_value_dict = data_source[row_key]
                    if not isinstance(row_value_dict, dict): continue # Skip invalid entries

                    row_dict = {}
                    # Map defined keys
                    for header_in_sheet, data_key in fob_data_keys.items():
                        target_col_idx = column_map.get(header_in_sheet)
                        if target_col_idx:
                            value = row_value_dict.get(data_key)
                            # Handle potential string numbers
                            if isinstance(value, str):
                                try:
                                    cleaned_val = value.replace(',', '').strip()
                                    if cleaned_val: # Avoid converting empty string
                                        if '.' in cleaned_val or 'e' in cleaned_val.lower(): value = float(cleaned_val)
                                        else: value = int(cleaned_val)
                                    else: value = None # Treat empty string as None/blank
                                except (ValueError, TypeError): pass # Keep as string if conversion fails
                            elif isinstance(value, Decimal): value = float(value)

                            row_dict[target_col_idx] = value

                    # Handle Description with fallback
                    if desc_col_idx_fob_prep:
                        desc_value_fob = row_value_dict.get('combined_description')
                        # Fallback if empty/None
                        if desc_value_fob is None or not str(desc_value_fob).strip():
                            desc_header_text = idx_to_header_map.get(desc_col_idx_fob_prep)
                            if desc_header_text:
                                for map_rule in mapping_rules.values():
                                    if isinstance(map_rule, dict) and map_rule.get('header') == desc_header_text:
                                        fallback = map_rule.get('fallback_on_none')
                                        if fallback is None: fallback = map_rule.get('static_value')
                                        if fallback is not None: desc_value_fob = fallback
                                        break # Found rule
                        row_dict[desc_col_idx_fob_prep] = desc_value_fob

                    # Add static values not already populated
                    for static_col_idx, static_val in static_value_map.items():
                         if static_col_idx not in row_dict: row_dict[static_col_idx] = static_val

                    data_rows_prepared.append(row_dict)
                print(f"DEBUG: FOB Mode - Prepared {len(data_rows_prepared)} data rows.")
            else:
                 print(f"Warning: FOB Mode - data_source is not a valid dictionary.")

            # Note: Pallet count logic needs review based on how FOB data relates to pallets.
            # For now, we'll assume 0 pallets per row for FOB in row-level calculations.
            pallet_counts_for_rows = [0] * len(data_rows_prepared)

        else:
            print(f"Error: Unknown data_source_type '{data_source_type}'")
            return False, data_writing_start_row, -1, -1, 0 # Use data_writing_start_row

        # --- Determine Final Number of Data Rows ---
        if data_source_type == 'fob_aggregation':
            # FOB mode: Number of prepared data rows + number of initial static labels
            # actual_rows_to_process = len(data_rows_prepared) + num_static_labels # OLD logic
            # Revised FOB logic: Process rows for the MAX of static labels or data rows
            actual_rows_to_process = max(len(data_rows_prepared), num_static_labels)
        else:
            # Standard modes: Based on prepared data or static labels
            total_data_rows_needed = max(len(data_rows_prepared), num_static_labels)
            actual_rows_to_process = total_data_rows_needed

        # Apply max_rows_to_fill constraint (only relevant for non-FOB modes typically)
        if max_rows_to_fill is not None and max_rows_to_fill >= 0: actual_rows_to_process = min(total_data_rows_needed, max_rows_to_fill)

        # Ensure pallet counts list matches the number of rows we intend to process
        if len(pallet_counts_for_rows) < actual_rows_to_process: pallet_counts_for_rows.extend([0] * (actual_rows_to_process - len(pallet_counts_for_rows)))
        elif len(pallet_counts_for_rows) > actual_rows_to_process: pallet_counts_for_rows = pallet_counts_for_rows[:actual_rows_to_process]

        # --- Calculate Total Rows to Insert and Row Indices ---
        total_rows_to_insert = 0
        current_row_offset = 0

        # Row after header (static/blank)
        if add_blank_after_header:
            row_after_header_idx = data_writing_start_row + current_row_offset
            total_rows_to_insert += 1
            current_row_offset += 1
        else:
            row_after_header_idx = -1 # Indicate no blank row

        # Data rows
        data_start_row = data_writing_start_row + current_row_offset
        if actual_rows_to_process > 0:
            data_end_row = data_start_row + actual_rows_to_process - 1
            total_rows_to_insert += actual_rows_to_process
            current_row_offset += actual_rows_to_process
        else:
            # No data rows to process (can happen if source is empty)
            data_end_row = data_start_row - 1 # Indicate no data rows

        # Row before footer (static/blank)
        if add_blank_before_footer:
            row_before_footer_idx = data_writing_start_row + current_row_offset
            total_rows_to_insert += 1
            current_row_offset += 1
        else:
            row_before_footer_idx = -1 # Indicate no blank row

        # Calculate final footer row index relative to where this chunk starts
        footer_row_final = data_writing_start_row + total_rows_to_insert
        total_rows_to_insert += 1 # Add 1 for the footer itself

        # --- Bulk Insert Rows --- # V11: Only insert if NOT pre-inserted by caller (i.e., for single-table modes)
        if data_source_type in ['aggregation', 'fob_aggregation']:
            if total_rows_to_insert > 0:
                try:
                    worksheet.insert_rows(data_writing_start_row, amount=total_rows_to_insert)
                    # Unmerge the block covering the inserted rows *before* the footer starts
                    safe_unmerge_block(worksheet, data_writing_start_row, footer_row_final - 1, num_columns)
                    print("Rows inserted and unmerged successfully.")
                except Exception as bulk_insert_err:
                    print(f"Error during single-table bulk row insert/unmerge: {bulk_insert_err}")
                    # Adjust fallback row calculation
                    fallback_row = max(header_info.get('second_row_index', 0) + 1, footer_row_final)
                    return False, fallback_row, -1, -1, 0

        # --- Fill Row After Header (if applicable) --- 

        # --- Prepare FOB Data Dictionary (inside loop now, safer) ---
        # Removed the premature preparation block here.
        # FOB data dict will be prepared inside the loop if data_source_type is fob_aggregation.

        # --- Fill Data Rows Loop ---
        if actual_rows_to_process > 0:
            print(f"--- DEBUG START LOOP (Sheet: {worksheet.title}) ---")
            print(f"  data_start_row: {data_start_row}")
            print(f"  actual_rows_to_process: {actual_rows_to_process}")
            print(f"  num_static_labels: {num_static_labels}")
            print(f"  col1_index: {col1_index}")
            print(f"  initial_static_col1_values: {initial_static_col1_values}")
            print(f"  data_source_type: {data_source_type}")
            # --- END DEBUG START LOOP ---
            try:
                headers_to_sum = ["Amount(USD)", "PCS", "SF", "N.W (kgs)", "G.W (kgs)", "CBM","Quantity(SF)", "Quantity\n(SF)", "Amount(USD)", "Quantity ( SF )", "Amount ( USD )", "Quantity(SF)", "Total value(USD)", "Quantity"]
                # Define general NO. column index needed for standard modes
                no_col_idx = column_map.get("NO") or column_map.get("NO.") # Moved definition here
                # Pre-find column indices needed for FOB mode
                fob_unit_price_col_idx = None
                fob_amount_col_idx = None
                fob_quantity_col_idx = None
                fob_no_col_idx = None
                fob_pallet_info_col_idx = pallet_info_col_idx # Reuse general pallet info index if found

                unit_price_headers = ["Unit price ( USD )", "Unit Price(USD)", "Unit Price\n(USD)", "Unit Price"]
                amount_headers = ["Amount ( USD )", "Total value(USD)", "Amount(USD)"]
                quantity_headers = ["Quantity ( SF )", "Quantity(SF)", "Quantity\n(SF)"]
                no_headers = ["NO", "NO."]

                for header, col_idx in column_map.items():
                    if header in unit_price_headers: fob_unit_price_col_idx = col_idx
                    if header in amount_headers: fob_amount_col_idx = col_idx
                    if header in quantity_headers: fob_quantity_col_idx = col_idx
                    if header in no_headers: fob_no_col_idx = col_idx
                    # fob_pallet_info_col_idx is already set from pallet_info_col_idx

                row_pallet_index = 0 # Counter for rows consuming pallets within this chunk

                for i in range(actual_rows_to_process):
                    target_row = data_start_row + i
                    data_row_indices_written.append(target_row)
                    # row_data_dict = data_rows_prepared[i] if i < len(data_rows_prepared) else {} # Old, not needed for FOB this way
                    is_last_data_row = (i == actual_rows_to_process - 1)

                    # --- Modified Pallet Count Logic ---
                    current_row_pallet_count = 0
                    try:
                        # Get the pallet count for the current row (only relevant for non-FOB modes currently)
                        if data_source_type != 'fob_aggregation':
                            raw_count = pallet_counts_for_rows[i] # Use index 'i' here
                            current_row_pallet_count = int(raw_count) if isinstance(raw_count, (int, float)) or (isinstance(raw_count, str) and raw_count.isdigit()) else 0
                            current_row_pallet_count = max(0, current_row_pallet_count)
                    except (IndexError, ValueError, TypeError):
                        pass # Keep current_row_pallet_count as 0 on error

                    # Increment index *if* this row consumes pallets
                    if current_row_pallet_count > 0:
                         row_pallet_index += 1

                    # Set the display value for the 'X' part using the row's pallet index within this chunk
                    display_pallet_order = row_pallet_index



                    # --- Cell Filling Logic --- #
                    if data_source_type == 'fob_aggregation':
                        # --- FOB Mode Filling (Revised - Process Row by Row) --- #
                        # print(f"DEBUG: FOB Mode - Processing row {target_row} (i={i})")

                        # 1. Write Static Label (if applicable for this row index)
                        if i < num_static_labels:
                            # print(f"DEBUG: FOB - Writing static label for i={i}")
                            if col1_index and i < len(initial_static_col1_values):
                                static_val_to_write = initial_static_col1_values[i]
                                if static_val_to_write is not None:
                                    try:
                                        cell_static = worksheet.cell(row=target_row, column=col1_index, value=static_val_to_write)
                                        _apply_cell_style(cell_static, static_column_header_name, sheet_styling_config)
                                        # Ensure static label col is text if needed (less common, but safety)
                                        if static_column_header_name in force_text_headers and cell_static.number_format != FORMAT_TEXT:
                                            cell_static.number_format = FORMAT_TEXT
                                    except Exception as static_write_err:
                                        print(f"Warning: FOB Error writing static value '{static_val_to_write}' to {target_row},{col1_index}: {static_write_err}")
                            else:
                                # This case means static label config exists but no col1_index found, or index out of bounds
                                pass
                                # print(f"DEBUG: FOB - Skipping static label for i={i} (col1_index: {col1_index}, len: {len(initial_static_col1_values)})")

                        # 2. Write Data (if applicable for this row index)
                        if i < len(data_rows_prepared):
                            current_fob_row_data = data_rows_prepared[i]
                            # print(f"DEBUG: FOB - Writing data for i={i} from: {current_fob_row_data}")

                            # Pre-calculate formula parts for this row
                            amount_col_letter = get_column_letter(fob_amount_col_idx) if fob_amount_col_idx else None
                            quantity_col_letter = get_column_letter(fob_quantity_col_idx) if fob_quantity_col_idx else None

                            for c_idx in range(1, num_columns + 1):
                                # --- Skip writing data to the static label column --- #
                                if i < num_static_labels and c_idx == col1_index:
                                    # print(f"DEBUG: FOB - Skipping data write to static col {c_idx} for i={i}")
                                    continue # Don't overwrite static label with data

                                cell = worksheet.cell(row=target_row, column=c_idx)
                                current_header = idx_to_header_map.get(c_idx)
                                value_to_write = None
                                is_force_text_column = current_header in force_text_headers

                                try:
                                    # --- Determine Value --- #
                                    if c_idx == fob_no_col_idx:
                                        value_to_write = i + 1 # 1-based row number for data index i
                                    elif c_idx == fob_unit_price_col_idx and amount_col_letter and quantity_col_letter:
                                        qty_cell_ref = f"{quantity_col_letter}{target_row}"
                                        amt_cell_ref = f"{amount_col_letter}{target_row}"
                                        value_to_write = f"={amt_cell_ref}/{qty_cell_ref}"
                                        # print(f"DEBUG: FOB - Set Unit Price formula {c_idx}: {value_to_write}")
                                    elif c_idx == fob_pallet_info_col_idx:
                                        # Use data row index i for pallet numbering within FOB group
                                        pallet_order_fob = i + 1
                                        value_to_write = f"{pallet_order_fob}-{local_chunk_pallets}"
                                        # print(f"DEBUG: FOB - Set Pallet Info {c_idx}: {value_to_write}")
                                    else:
                                        # Get value from prepared data for this index i
                                        value_to_write = current_fob_row_data.get(c_idx)
                                        # Type conversion already handled during preparation

                                    # --- Write Value --- #
                                    cell.value = value_to_write

                                    # --- Apply Formatting and Styles --- #
                                    if is_force_text_column and cell.number_format != FORMAT_TEXT:
                                        cell.number_format = FORMAT_TEXT
                                    elif c_idx == fob_pallet_info_col_idx and cell.number_format != FORMAT_TEXT:
                                         cell.number_format = FORMAT_TEXT

                                    # Apply general cell style
                                    _apply_cell_style(cell, current_header, sheet_styling_config)

                                except Exception as write_err:
                                    print(f"Warning: FOB - Error writing value '{value_to_write}' to {cell.coordinate}: {write_err}")
                                    cell.value = "#WRITE_ERR!" # Indicate error in cell
                        # else:
                            # print(f"DEBUG: FOB - No data to write for i={i}")

                        # 3. Ensure remaining columns are blank if only static label was written
                        if i < num_static_labels and i >= len(data_rows_prepared):
                            # print(f"DEBUG: FOB - Blanking remaining cells for static-only row i={i}")
                            for c_idx in range(1, num_columns + 1):
                                if c_idx != col1_index:
                                    try: worksheet.cell(row=target_row, column=c_idx).value = None
                                    except: pass # Ignore errors blanking cells

                    else: # --- Standard Mode Filling (Remains unchanged) --- #
                        # Determine row type flags ONCE per row
                        row_data_dict = data_rows_prepared[i] if i < len(data_rows_prepared) else {}
                        is_data_row = (i < len(data_rows_prepared))
                        is_static_label_row = (i < num_static_labels)
                    # merging rule for data cells

                        for c_idx in range(1, num_columns + 1):
                            cell = worksheet.cell(row=target_row, column=c_idx)
                            current_header = idx_to_header_map.get(c_idx)
                            value_to_write = None # Default value
                            is_force_text_column = current_header in force_text_headers

                            # --- Priority 1: Handle Static Label Column --- #
                            if is_static_label_row and c_idx == col1_index:
                                static_val_to_write = initial_static_col1_values[i] if i < len(initial_static_col1_values) else None
                                value_to_write = static_val_to_write
                                try:
                                    cell.value = value_to_write
                                    _apply_cell_style(cell, static_column_header_name, sheet_styling_config)
                                except Exception as static_write_err:
                                     print(f"Warning: Error writing static value '{static_val_to_write}' to {target_row},{col1_index}: {static_write_err}")
                                continue # Static label takes precedence for this cell, move to next column

                            # --- Priority 2: Handle Data Rows (for all columns other than the static label column handled above) --- #
                            if is_data_row:
                                # row_data_dict already fetched above

                                # --- Determine if Custom Overrides apply to THIS cell (c_idx) --- #
                                custom_unit_price_formula = None
                                skip_amount_config_formula = False

                                if custom_flag and data_source_type == 'aggregation':
                                    # Check for Unit Price Column
                                    unit_price_col_idx_formula = None
                                    unit_price_headers_formula = ["Unit price ( USD )", "Unit Price(USD)", "unit price", "Unit Price\n(USD)"]
                                    for header, col_idx in column_map.items():
                                        if str(header).lower() in [h.lower() for h in unit_price_headers_formula]:
                                            unit_price_col_idx_formula = col_idx
                                            break
                                    if c_idx == unit_price_col_idx_formula:
                                        # Calculate custom formula for Unit Price
                                        amount_col_idx_formula = None
                                        quantity_col_idx_formula = None
                                        amount_headers_formula = ["Amount ( USD )", "Total value(USD)", "amount", "amount_sum", "Amount(USD)"]
                                        quantity_headers_formula = ["Quantity ( SF )", "Quantity(SF)", "Quantity", "sqft", "sqft_sum", "Quantity\n(SF)"]
                                        for header, col_idx in column_map.items():
                                            header_lower = str(header).lower()
                                            if header_lower in [h.lower() for h in amount_headers_formula]: amount_col_idx_formula = col_idx
                                            if header_lower in [h.lower() for h in quantity_headers_formula]: quantity_col_idx_formula = col_idx
                                        if amount_col_idx_formula and quantity_col_idx_formula:
                                            try:
                                                amount_col_letter = get_column_letter(amount_col_idx_formula)
                                                quantity_col_letter = get_column_letter(quantity_col_idx_formula)
                                                custom_unit_price_formula = f"={amount_col_letter}{target_row}/{quantity_col_letter}{target_row}"
                                            except Exception: custom_unit_price_formula = "#FORMULA_ERR!"
                                        else: custom_unit_price_formula = "#REF!"
                                        print(f"DEBUG: Custom Aggregation - Setting Unit Price Formula for {target_row},{c_idx}: {custom_unit_price_formula}")

                                    # Check for Amount Column (to skip config formula later)
                                    amount_col_idx_override_check = None
                                    amount_headers_override_check = ["Amount ( USD )", "Total value(USD)", "amount", "amount_sum", "Amount(USD)"]
                                    for header, col_idx in column_map.items():
                                        if str(header).lower() in [h.lower() for h in amount_headers_override_check]:
                                            amount_col_idx_override_check = col_idx
                                            break
                                    if c_idx == amount_col_idx_override_check:
                                        skip_amount_config_formula = True
                                        print(f"DEBUG: Custom Aggregation - Identified Amount column {c_idx}. Config formula will be skipped.")

                                # --- Determine Value to Write for Data Cell --- #
                                if custom_unit_price_formula is not None:
                                    value_to_write = custom_unit_price_formula # Use the custom formula
                                # --- Add Pallet Info Check Here --- #
                                elif pallet_info_col_idx is not None and c_idx == pallet_info_col_idx:
                                    # Use the locally calculated index for 'X'
                                    value_to_write = f"{display_pallet_order}-{local_chunk_pallets}"
                                # --- End Pallet Info Check --- #
                                # --- Add NO. Column Check Here (Standard Mode) --- #
                                elif no_col_idx is not None and c_idx == no_col_idx: # Use the general no_col_idx found earlier
                                    value_to_write = i + 1 # Use loop index (0-based) + 1 for sequential numbering
                                # --- End NO. Column Check --- #
                                elif c_idx in formula_rules and not skip_amount_config_formula: # Apply config formula (if not skipped)
                                    print(f"DEBUG: Applying config formula for column {c_idx}")
                                    rule = formula_rules[c_idx]; formula_template = rule["template"]; input_headers = rule["input_headers"]
                                    formula_params = {'row': target_row}; valid_inputs = True
                                    for idx, input_header in enumerate(input_headers):
                                        input_col_idx = column_map.get(input_header)
                                        if input_col_idx: formula_params[f'col_ref_{idx}'] = get_column_letter(input_col_idx)
                                        else: valid_inputs = False; break
                                    if valid_inputs:
                                        try: value_to_write = f"={formula_template.format(**formula_params)}"
                                        except Exception: value_to_write = "#ERR!"
                                    else: value_to_write = "#REF!"
                                else: # Apply standard mapping (if no formula applied)
                                    value_to_write = row_data_dict.get(c_idx)
                                    # Handle type conversions for mapped values
                                    if isinstance(value_to_write, str):
                                        try:
                                            cleaned_str = value_to_write.replace(',', '').strip()
                                            if cleaned_str: value_to_write = float(cleaned_str) if '.' in cleaned_str or 'e' in cleaned_str.lower() else int(cleaned_str)
                                            elif not (current_header in headers_to_sum): pass # Keep non-numeric strings if not a sum col
                                            else: value_to_write = None # Blank out sum columns if empty string
                                        except ValueError: pass # Keep as string if conversion fails
                                    elif isinstance(value_to_write, Decimal): value_to_write = float(value_to_write)

                                # --- Write Value and Apply Style for Data Cell --- #
                                cell.value = value_to_write
                                is_force_text_column = current_header in force_text_headers
                                if is_force_text_column and cell.number_format != FORMAT_TEXT: cell.number_format = FORMAT_TEXT
                                elif pallet_info_col_idx == c_idx and cell.number_format != FORMAT_TEXT: cell.number_format = FORMAT_TEXT
                                _apply_cell_style(cell, current_header, sheet_styling_config)

                            # --- Priority 3: Handle Purely Static Rows (non-data), Non-Label Columns --- #
                            elif is_static_label_row: # Only executes if is_data_row was false AND is_static_label_row is true
                                # We already handled c_idx == col1_index with 'continue'
                                # So this handles other columns in rows that ONLY have a static label
                                cell.value = None # Ensure blank
                                # Apply minimal styling if needed (e.g., borders)
                                # _apply_cell_style(cell, current_header, sheet_styling_config) # Optional styling for blank static cells
                            # --- Else (row is neither data nor static label - should not happen if actual_rows_to_process is correct) --- #
                            # Optionally handle this case if needed, but likely implies row count mismatch

                    # --- Apply Border (Common to both modes, done once per row) ---
                    for c_idx_border in range(1, num_columns + 1):
                        try: 
                            cell = worksheet.cell(row=target_row, column=c_idx_border)
                            current_header = idx_to_header_map.get(c_idx_border)
                            col_styles = sheet_styling_config.get("column_styles", {}) if sheet_styling_config else {}
                            col_specific_style = col_styles.get(current_header, {}) if current_header else {}
                            apply_grid_col = col_specific_style.get("border") == "full_grid"
                            apply_grid_list = current_header and current_header in columns_to_grid
                            apply_grid = apply_grid_col or apply_grid_list
                            top_b = thin_side if i == 0 else (thin_side if apply_grid else None)
                            bottom_b = thin_side if is_last_data_row else (thin_side if apply_grid else None)
                            is_initial_static_col_border = (c_idx_border == col1_index)
                            if apply_special_border_rule and is_initial_static_col_border: cell.border = Border(left=thin_side, right=thin_side, top=(thin_side if i==0 else None), bottom=None)
                            elif apply_grid: cell.border = thin_border
                            else: cell.border = Border(left=thin_side, right=thin_side, top=top_b, bottom=bottom_b)
                        except IndexError: 
                            continue # Skip if cell index is somehow out of bounds
                        except Exception as border_err:
                            # Add an except block to catch other potential border errors
                            print(f"Warning: Error applying border to {target_row},{c_idx_border}: {border_err}")
                            pass # Continue processing other cells/rows

                    if data_cell_merging_rules: # Check if there are any rules for this sheet
                        apply_explicit_data_cell_merges(
                                worksheet=worksheet,
                                row_num=target_row,
                                column_map=column_map,
                                num_total_columns=num_columns,
                                merge_rules_data_cells=data_cell_merging_rules, # Your variable name
                                sheet_styling_config=sheet_styling_config
                            )
            except Exception as fill_data_err:
                print(f"Error during data filling loop: {fill_data_err}\n{traceback.format_exc()}")
                return False, footer_row_final + 1, data_start_row, data_end_row, 0

            # --- Merge Description Column ---
            # Revised logic (v6): Merge contiguous cells containing actual fallback text in the description column
            # ONLY IF the description column was populated primarily by fallbacks/static values (i.e., dynamic_desc_used is False).
            # If descriptions came from at least some dynamic source data (dynamic_desc_used is True), skip this merge.
            # Also, if dynamic_desc_used is False but the cells are genuinely empty (None/whitespace), do not merge them.

            # The dynamic_desc_used flag is set earlier in the fill_invoice_data function.
            # It's True if any row's description came from direct, non-empty source data.
            # It's False if all row descriptions are from fallbacks, static values, or remained empty.

            if desc_col_idx is not None and actual_rows_to_process > 1 and not dynamic_desc_used: # Fallback scenario
                
                # --- Actual Description Merge Logic Starts Here (v6 - Merge non-empty fallback text ONLY IF dynamic_desc_used is False) ---
                # This code runs if desc_col_idx is not None, actual_rows_to_process > 1, AND dynamic_desc_used is False.

                # Use the globally defined center_alignment for merged description cells.
                merged_desc_alignment = center_alignment # Alignment(horizontal='center', vertical='center', wrap_text=True)

                current_merge_block_start_idx = data_start_row
                value_to_match_in_block = None
                try:
                    value_to_match_in_block = worksheet.cell(row=current_merge_block_start_idx, column=desc_col_idx).value
                except Exception as e_cell_access:
                    print(f"ERROR: Could not access initial cell for description merge at row {current_merge_block_start_idx}, col {desc_col_idx}. Error: {e_cell_access}")
                    value_to_match_in_block = object() # Unique object to prevent matching

                # Determine if the value of the current block being tracked is considered "empty"
                is_current_block_value_empty = False
                if value_to_match_in_block is None:
                    is_current_block_value_empty = True
                elif isinstance(value_to_match_in_block, str) and not value_to_match_in_block.strip():
                    is_current_block_value_empty = True
                
                for row_idx_for_comparison in range(data_start_row + 1, data_end_row + 2):
                    current_cell_value_for_comparison = None 
                    is_iteration_past_data_end = (row_idx_for_comparison > data_end_row)

                    if not is_iteration_past_data_end:
                        try:
                            current_cell_value_for_comparison = worksheet.cell(row=row_idx_for_comparison, column=desc_col_idx).value
                        except Exception as e_cell_access_loop:
                            print(f"ERROR: Could not access cell for description merge at row {row_idx_for_comparison}, col {desc_col_idx}. Error: {e_cell_access_loop}")
                            current_cell_value_for_comparison = object() # Unique object

                    if current_cell_value_for_comparison != value_to_match_in_block or is_iteration_past_data_end:
                        block_end_row_idx = row_idx_for_comparison - 1
                        
                        # Only perform the merge if:
                        # 1. The block spans more than one row.
                        # 2. The value of the block itself WAS NOT considered empty (i.e., it's actual fallback text).
                        if block_end_row_idx > current_merge_block_start_idx and not is_current_block_value_empty: # <<<< MODIFIED CONDITION HERE
                            try:
                                worksheet.merge_cells(
                                    start_row=current_merge_block_start_idx,
                                    start_column=desc_col_idx,
                                    end_row=block_end_row_idx,
                                    end_column=desc_col_idx
                                )
                                top_cell_of_merged_block = worksheet.cell(row=current_merge_block_start_idx, column=desc_col_idx)
                                top_cell_of_merged_block.alignment = merged_desc_alignment
                                
                            except Exception as e_merge:
                                print(f"WARNING: Could not merge description cells from {current_merge_block_start_idx} to {block_end_row_idx}. Error: {e_merge}")
                        elif block_end_row_idx > current_merge_block_start_idx and is_current_block_value_empty:
                            # This case (empty block) is now skipped even if dynamic_desc_used is False.
                            print(f"DEBUG: Skipped merging EMPTY description block in column {desc_col_idx} from row {current_merge_block_start_idx} to {block_end_row_idx} (value was None or whitespace).")
                            pass 

                        if not is_iteration_past_data_end:
                            current_merge_block_start_idx = row_idx_for_comparison
                            value_to_match_in_block = current_cell_value_for_comparison
                            
                            # Update emptiness check for the new block
                            if value_to_match_in_block is None:
                                is_current_block_value_empty = True
                            elif isinstance(value_to_match_in_block, str) and not value_to_match_in_block.strip():
                                is_current_block_value_empty = True
                            else:
                                is_current_block_value_empty = False
                # --- Actual Description Merge Logic Ends Here (v6) ---
                print(f":::::::::::::::::::: Finished Description cell merge attempt for sheet '{sheet_name}' (occurred because dynamic_desc_used was False)") # DEBUG
            else:
                # Print why merge is skipped
                if desc_col_idx is None:
                    print(f":::::::::::::::::::: Skipping Description merge for sheet '{sheet_name}' - Description column not found.")
                elif actual_rows_to_process <= 1:
                    print(f":::::::::::::::::::: Skipping Description merge for sheet '{sheet_name}' - Not enough rows ({actual_rows_to_process})")
                elif dynamic_desc_used: # Check this condition (dynamic_desc_used is True)
                    print(f":::::::::::::::::::: Skipping Description merge for sheet '{sheet_name}' - Descriptions used dynamic source data (dynamic_desc_used is True). Fallback merge logic is bypassed.")

# --- Merge Pallet Info Column (if applicable, similar to Description) ---
            # This block should be placed AFTER the "Merge Description Column" block.
            # It merges contiguous cells in the "Pallet\nNo" column if they have the same non-empty value,
            # and if dynamic_desc_used is False (implying a more static/repetitive layout).
            # get pallet_info_index

            if pallet_info_col_idx is not None and actual_rows_to_process > 1:
                print(f":::::::::::::::::::: Starting Pallet Info cell merge for sheet '{sheet_name}' (dynamic_desc_used is False).")

                # Use the same center_alignment as used for description or define one.
                # merged_pallet_alignment = center_alignment

                current_merge_block_start_idx_pallet = data_start_row
                value_to_match_in_block_pallet = None
                try:
                    value_to_match_in_block_pallet = worksheet.cell(row=current_merge_block_start_idx_pallet, column=pallet_info_col_idx).value
                except Exception as e_cell_access_pallet:
                    print(f"ERROR: Could not access initial cell for pallet info merge at row {current_merge_block_start_idx_pallet}, col {pallet_info_col_idx}. Error: {e_cell_access_pallet}")
                    value_to_match_in_block_pallet = object() # Unique object to prevent matching

                # Determine if the value of the current block being tracked is considered "empty"
                is_current_block_value_empty_pallet = False
                if value_to_match_in_block_pallet is None:
                    is_current_block_value_empty_pallet = True
                elif isinstance(value_to_match_in_block_pallet, str) and not str(value_to_match_in_block_pallet).strip():
                    is_current_block_value_empty_pallet = True

                for row_idx_for_comparison_pallet in range(data_start_row + 1, data_end_row + 2): # Iterate one past the end
                    current_cell_value_for_comparison_pallet = None
                    is_iteration_past_data_end_pallet = (row_idx_for_comparison_pallet > data_end_row)

                    if not is_iteration_past_data_end_pallet:
                        try:
                            current_cell_value_for_comparison_pallet = worksheet.cell(row=row_idx_for_comparison_pallet, column=pallet_info_col_idx).value
                        except Exception as e_cell_access_loop_pallet:
                            print(f"ERROR: Could not access cell for pallet info merge at row {row_idx_for_comparison_pallet}, col {pallet_info_col_idx}. Error: {e_cell_access_loop_pallet}")
                            current_cell_value_for_comparison_pallet = object() # Unique object

                    # If value changes or we are past the last data row, finalize the previous block
                    if current_cell_value_for_comparison_pallet != value_to_match_in_block_pallet or is_iteration_past_data_end_pallet:
                        block_end_row_idx_pallet = row_idx_for_comparison_pallet - 1

                        # Only perform the merge if:
                        # 1. The block spans more than one row.
                        # 2. The value of the block itself WAS NOT considered empty.
                        if block_end_row_idx_pallet > current_merge_block_start_idx_pallet and not is_current_block_value_empty_pallet:
                            try:
                                worksheet.merge_cells(
                                    start_row=current_merge_block_start_idx_pallet,
                                    start_column=pallet_info_col_idx,
                                    end_row=block_end_row_idx_pallet,
                                    end_column=pallet_info_col_idx
                                )
                                top_cell_of_merged_block_pallet = worksheet.cell(row=current_merge_block_start_idx_pallet, column=pallet_info_col_idx)
                                top_cell_of_merged_block_pallet.alignment = center_alignment # Or merged_pallet_alignment
                                
                            except Exception as e_merge_pallet:
                                print(f"WARNING: Could not merge Pallet Info cells from {current_merge_block_start_idx_pallet} to {block_end_row_idx_pallet}. Error: {e_merge_pallet}")
                        elif block_end_row_idx_pallet > current_merge_block_start_idx_pallet and is_current_block_value_empty_pallet:
                            print(f"DEBUG: Skipped merging EMPTY Pallet Info block in column {pallet_info_col_idx} from row {current_merge_block_start_idx_pallet} to {block_end_row_idx_pallet} (value was None or whitespace).")
                            pass

                        # Start a new block
                        if not is_iteration_past_data_end_pallet:
                            current_merge_block_start_idx_pallet = row_idx_for_comparison_pallet
                            value_to_match_in_block_pallet = current_cell_value_for_comparison_pallet
                            
                            # Update emptiness check for the new block
                            if value_to_match_in_block_pallet is None:
                                is_current_block_value_empty_pallet = True
                            elif isinstance(value_to_match_in_block_pallet, str) and not str(value_to_match_in_block_pallet).strip():
                                is_current_block_value_empty_pallet = True
                            else:
                                is_current_block_value_empty_pallet = False
                print(f":::::::::::::::::::: Finished Pallet Info cell merge attempt for sheet '{sheet_name}'.")
            else:
                # Print why pallet merge is skipped
                if pallet_info_col_idx is None:
                    print(f":::::::::::::::::::: Skipping Pallet Info merge for sheet '{sheet_name}' - Pallet Info column not found.")
                elif actual_rows_to_process <= 1:
                    print(f":::::::::::::::::::: Skipping Pallet Info merge for sheet '{sheet_name}' - Not enough rows ({actual_rows_to_process}).")
                elif dynamic_desc_used: # Check this condition
                    print(f":::::::::::::::::::: Skipping Pallet Info merge for sheet '{sheet_name}' - Dynamic descriptions were used (dynamic_desc_used is True), so pallet merge is also skipped.")
            # --- End Merge Pallet Info Column ---



# --- Fill Row Before Footer ---
        if add_blank_before_footer and row_before_footer_idx > 0:
            try:
                 fill_static_row(worksheet, row_before_footer_idx, num_columns, static_content_before_footer) # Applies no_border by default
                 # Re-apply styles and specific borders for this row
                 for c_idx in range(1, num_columns + 1):
                     cell = worksheet.cell(row=row_before_footer_idx, column=c_idx)
                     # Apply general cell styling (font, alignment, number format) first
                     _apply_cell_style(cell, idx_to_header_map.get(c_idx), sheet_styling_config) 
                     
                     # Now, apply specific border logic for this row
                     if c_idx == col1_index: 
                        # For col1_index (e.g., "Mark & Nº"), only left and right borders, no top/bottom
                        cell.border = Border(left=thin_side, right=thin_side, top=None, bottom=None)
                     else: 
                        # For all other columns in the "row before footer", apply full thin borders
                        cell.border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
            except Exception as fill_bf_err: print(f"Warning: Error filling/styling row before footer: {fill_bf_err}")

        # --- Fill Footer Row --- (Keep existing logic)
        # The SUM formulas here should correctly sum the results of the formulas
        # written in the data rows above.
        total_value_config = sheet_config.get("total_footer_text")
        if footer_row_final > 0:
             # <<< Add this line >>>
             unmerge_row(worksheet, footer_row_final, num_columns) # Ensure footer row is clear before writing
             try:
                 palletNo_col_inx = None
                 headers_to_sum = ["PCS", "SF", "N.W (kgs)", "G.W (kgs)", "CBM", "Quantity ( SF )", "Amount ( USD )", "Quantity\n(SF)", "Total value(USD)", "Quantity", "Quantity(SF)", "Amount(USD)"]
                 total_text_col_idx = None; item_col_idx = column_map.get("NO") or column_map.get("NO."); total_text_col_idx = item_col_idx
                 if not total_text_col_idx: palletNo_col_inx = column_map.get("PALLET\nNO.") or column_map.get("Pallet\nNo"); total_text_col_idx = palletNo_col_inx
                 if not total_text_col_idx: po_col_idx = column_map.get("P.O Nº") or column_map.get("P.O N°") or column_map.get("CUT.P.O.") ; total_text_col_idx = po_col_idx
                 if not total_text_col_idx: mark_col_idx = column_map.get("Mark & Nº") or column_map.get("Mark & N °"); total_text_col_idx = mark_col_idx
                 if not total_text_col_idx: desc_col_idx = column_map.get("placeholder") or column_map.get("Description\nOf Goods") or column_map.get("Description of Goods"); total_text_col_idx = desc_col_idx
                 if not total_text_col_idx: total_text_col_idx = desc_col_idx if desc_col_idx else 1
                 if total_text_col_idx:
                     try: total_cell = worksheet.cell(row=footer_row_final, column=total_text_col_idx, value= total_value_config or "TOTAL OF: "); total_cell.font = effective_header_font; total_cell.alignment = effective_header_align # Removed _apply_cell_style call
                     except Exception as e: print(f"Warning: Error writing 'TOTAL OF:' text: {e}")

                 sum_range_valid = actual_rows_to_process > 0 and data_start_row <= data_end_row
                 sum_start_row_for_formula = data_start_row; sum_end_row_for_formula = data_end_row
                 for header_name in headers_to_sum:
                     col_idx = column_map.get(header_name)
                     if col_idx:
                         formula_or_value = 0
                         if sum_range_valid: col_letter = get_column_letter(col_idx); formula_or_value = f"=SUM({col_letter}{sum_start_row_for_formula}:{col_letter}{sum_end_row_for_formula})"
                         try: cell = worksheet.cell(row=footer_row_final, column=col_idx, value=formula_or_value); cell.font = effective_header_font; cell.alignment = effective_header_align # Removed _apply_cell_style
                         except Exception as e: print(f"Warning: Error writing SUM formula for {header_name}: {e}")

                 # --- Footer Pallet Count --- (Configurable Column) ---
                 # Determine which pallet count to display (logic remains the same)
                 pallets_to_display = -1 # Default to error/skip state
                 if data_source_type == 'processed_tables': #
                     pallets_to_display = local_chunk_pallets # Use count local to this table
                 elif data_source_type in ['aggregation', 'fob_aggregation']: #
                     pallets_to_display = grand_total_pallets # Use the overall grand total passed in
                 else: #
                     print(f"Warning: Unknown data_source_type '{data_source_type}' for pallet footer.") #

                 # Determine the target column for the footer pallet count from config
                 footer_pallet_header = sheet_config.get("footer_pallet_count_column_header") # Get header name from config
                 footer_pallet_col_idx = None # Initialize column index

                 if footer_pallet_header: # If a header name was provided in config
                     footer_pallet_col_idx = column_map.get(footer_pallet_header) # Find its index
                     if footer_pallet_col_idx is None: # If header from config wasn't found in the sheet
                         print(f"Warning: Configured footer pallet count header '{footer_pallet_header}' not found. Falling back to description column.")
                         footer_pallet_col_idx = desc_col_idx # Fallback 1: Use Description column index
                 else:
                     # No specific config provided, default to Description column index
                     footer_pallet_col_idx = desc_col_idx

                 # Write the pallet count if the count is valid AND we found a valid column index
                 if footer_pallet_col_idx is not None and pallets_to_display >= 0: #
                     try:
                         # Use the determined footer_pallet_col_idx
                         pallet_cell = worksheet.cell(row=footer_row_final, column=footer_pallet_col_idx) # Use the determined index
                         pallet_cell.value = f"{pallets_to_display} PALLETS" #
                         pallet_cell.font = effective_header_font #
                         pallet_cell.alignment = effective_header_align #
                     except Exception as e: #
                         print(f"Warning: Error writing pallet count '{pallets_to_display}' to footer column {footer_pallet_col_idx}: {e}") # Updated warning
                 elif footer_pallet_col_idx is None: # Only print warning if column index is truly invalid
                     print(f"Warning: Cannot determine column for footer pallet count (Description column likely missing). Skipping.")
                 # else: # Handles pallets_to_display < 0 (error calculating count)
                     # Optionally clear the cell if needed
                     # try:
                     #     if footer_pallet_col_idx: worksheet.cell(row=footer_row_final, column=footer_pallet_col_idx).value = None
                     # except: pass

                 # --- Footer Border/Style --- 
                 for c_idx in range(1, num_columns + 1):
                     try:
                         cell = worksheet.cell(row=footer_row_final, column=c_idx); 
                         footer_header = idx_to_header_map.get(c_idx) 
                         apply_grid_footer = footer_header and footer_header in columns_to_grid 
                         top_border_side = thin_side; bottom_border_side = thin_side
                         
                         # Apply border first
                         if c_idx == col1_index: 
                             cell.border = Border(left=thin_side, right=thin_side, top=None, bottom=bottom_border_side)
                         elif apply_grid_footer: 
                             cell.border = thin_border
                         else: 
                             cell.border = Border(left=thin_side, right=thin_side, top=top_border_side, bottom=bottom_border_side)
                         
                         # Apply style MANUALLY for footer to preserve bold font but get number format
                         if cell.value is not None:
                             cell.font = effective_header_font 
                             cell.alignment = effective_header_align 
                             
                             # Manually get and apply number format from config
                             footer_number_format = None
                             if sheet_styling_config and footer_header:
                                 column_styles_cfg = sheet_styling_config.get("column_styles", {})
                                 col_specific_style_cfg = column_styles_cfg.get(footer_header, {})
                                 footer_number_format = col_specific_style_cfg.get("number_format")
                             
                             if footer_number_format and cell.number_format != FORMAT_TEXT:
                                 try: cell.number_format = footer_number_format
                                 except Exception: pass # Ignore errors applying format
                             # Don't call _apply_cell_style here to avoid overwriting font/alignment
                             
                     except Exception as e: print(f"Warning: Error styling footer cell {c_idx}: {e}")
             except Exception as fill_footer_err: print(f"Error during footer filling: {fill_footer_err}")

        # --- Apply Merges ---
        # Apply merges to row after header (if applicable)
        if add_blank_after_header and row_after_header_idx > 0 and merge_rules_after_header:
            apply_row_merges(worksheet, row_after_header_idx, num_columns, merge_rules_after_header)

        # Apply merges to row before footer (if applicable)
        target_row_for_bf_merge = row_before_footer_idx if add_blank_before_footer and row_before_footer_idx > 0 else -1
        if target_row_for_bf_merge > 0 and merge_rules_before_footer:
            apply_row_merges(worksheet, target_row_for_bf_merge, num_columns, merge_rules_before_footer)

        # Apply merges to the footer row itself (if applicable)
        if footer_row_final > 0 and merge_rules_footer:
            print(f"Applying footer merges to row {footer_row_final} with rules: {merge_rules_footer}") # Optional Debug
            try:
                apply_row_merges(worksheet, footer_row_final, num_columns, merge_rules_footer)
            except Exception as footer_merge_err:
                 print(f"Warning: Error applying footer merges: {footer_merge_err}")

        # --- Apply Row Heights --- (Keep existing)
        apply_row_heights(worksheet=worksheet, sheet_styling_config=sheet_styling_config, header_info=header_info, data_row_indices=data_row_indices_written, footer_row_index=footer_row_final, row_after_header_idx=row_after_header_idx, row_before_footer_idx=row_before_footer_idx)

        # --- Finalization --- (Keep existing)
        next_available_row_final = footer_row_final + 1
        if actual_rows_to_process == 0: data_start_row = -1; data_end_row = -1
        return True, next_available_row_final, data_start_row, data_end_row, local_chunk_pallets

    except Exception as e:
        # --- Error Handling --- (Keep existing)
        print(f"Critical error in fill_invoice_data: {e}\n{traceback.format_exc()}")
        fallback_row = header_info.get('second_row_index', 0) + 1; frf_local = locals().get('footer_row_final', -1)
        if frf_local > 0: fallback_row = max(fallback_row, frf_local + 1)
        else: est_footer = locals().get('initial_insert_point', fallback_row) + locals().get('total_rows_to_insert', 0); fallback_row = max(fallback_row, est_footer)
        return False, fallback_row, -1, -1, 0


def find_cell_by_marker(worksheet: Worksheet, marker_text: str, search_range: Optional[str] = None) -> Optional[openpyxl.cell.Cell]:
    """
    Finds the first cell containing the exact marker text within a specified range or the entire sheet.

    Args:
        worksheet: The openpyxl Worksheet object.
        marker_text: The exact string to search for in cell values.
        search_range: Optional string defining the range (e.g., "A1:F10"). Searches entire sheet if None.

    Returns:
        The openpyxl Cell object if found, otherwise None.
    """
    if not marker_text: return None
    cells_to_scan = None; search_description = "entire sheet"
    if search_range:
        try: cells_to_scan = worksheet[search_range]; search_description = f"range '{search_range}'"
        except Exception as range_err: cells_to_scan = worksheet.iter_rows() # Fallback to full scan on range error
    else: cells_to_scan = worksheet.iter_rows() # Default to full scan
    marker_text_str = str(marker_text).strip() # Ensure marker is string and stripped

    # Handle case where search_range might be a single cell
    if isinstance(cells_to_scan, openpyxl.cell.Cell):
        if cells_to_scan.value is not None and str(cells_to_scan.value).strip() == marker_text_str:
            return cells_to_scan
        else:
            return None # Single cell didn't match

    # Iterate through rows/cells
    if cells_to_scan is not None:
        try:
            for row in cells_to_scan:
                # Handle case where cells_to_scan is directly a tuple of cells (from worksheet[range])
                if isinstance(row, openpyxl.cell.Cell):
                    cell = row; # Treat the item itself as the cell
                    # Skip merged cells unless it's the top-left origin
                    if isinstance(cell, openpyxl.cell.cell.MergedCell): continue
                    if cell.value is not None and str(cell.value).strip() == marker_text_str: return cell
                else: # Assume row is an iterable of cells (from iter_rows)
                    for cell in row:
                        # Skip merged cells unless it's the top-left origin
                        if isinstance(cell, openpyxl.cell.cell.MergedCell): continue
                        if cell.value is not None and str(cell.value).strip() == marker_text_str: return cell
        except Exception as iter_err:
            # Log error if needed
            return None
    return None # Marker not found


def apply_column_widths(worksheet: Worksheet, sheet_styling_config: Optional[Dict[str, Any]], header_map: Optional[Dict[str, int]]):
    """
    Sets column widths based on the configuration.

    Args:
        worksheet: The openpyxl Worksheet object.
        sheet_styling_config: Styling configuration containing the 'column_widths' dictionary.
        header_map: Dictionary mapping header text to column index (1-based).
    """
    if not sheet_styling_config or not header_map: return
    column_widths_cfg = sheet_styling_config.get("column_widths")
    if not column_widths_cfg or not isinstance(column_widths_cfg, dict): return
    for header_text, width in column_widths_cfg.items():
        col_idx = header_map.get(header_text)
        if col_idx:
            col_letter = get_column_letter(col_idx)
            try:
                width_val = float(width)
                if width_val > 0: worksheet.column_dimensions[col_letter].width = width_val
                else: pass # Ignore non-positive widths
            except (ValueError, TypeError): pass # Ignore invalid width values
            except Exception as width_err: pass # Log other errors?
        else: pass # Header text not found in map


def apply_row_heights(worksheet: Worksheet, sheet_styling_config: Optional[Dict[str, Any]], header_info: Dict[str, Any], data_row_indices: List[int], footer_row_index: int, row_after_header_idx: int, row_before_footer_idx: int):
    """
    Sets row heights based on the configuration for header, data, footer, and specific rows.
    Footer height can now optionally match the header height.

    Args:
        worksheet: The openpyxl Worksheet object.
        sheet_styling_config: Styling configuration containing the 'row_heights' dictionary.
        header_info: Dictionary with header row indices.
        data_row_indices: List of 1-based indices for the actual data rows written.
        footer_row_index: 1-based index of the footer row.
        row_after_header_idx: 1-based index of the static/blank row after the header (-1 if none).
        row_before_footer_idx: 1-based index of the static/blank row before the footer (-1 if none).
    """
    if not sheet_styling_config: return
    row_heights_cfg = sheet_styling_config.get("row_heights")
    if not row_heights_cfg or not isinstance(row_heights_cfg, dict): return

    actual_header_height = None # Store the applied header height

    def set_height(r_idx, height_val, desc): # Helper function
        nonlocal actual_header_height # Ensure actual_header_height is modified
        if r_idx <= 0: return
        try:
            h_val = float(height_val)
            if h_val > 0:
                worksheet.row_dimensions[r_idx].height = h_val
                if desc == "header": # Store the height applied to the header
                    actual_header_height = h_val
            else: pass # Ignore non-positive heights
        except (ValueError, TypeError): pass # Ignore invalid height values
        except Exception as height_err: pass # Log other errors?

    # Apply Heights Based on Config
    header_height = row_heights_cfg.get("header")
    if header_height is not None and header_info:
        h_start = header_info.get('first_row_index', -1); h_end = header_info.get('second_row_index', -1)
        if h_start > 0 and h_end >= h_start:
            for r in range(h_start, h_end + 1): set_height(r, header_height, "header")

    after_header_height = row_heights_cfg.get("after_header")
    if after_header_height is not None and row_after_header_idx > 0: set_height(row_after_header_idx, after_header_height, "after_header")
    data_default_height = row_heights_cfg.get("data_default")
    if data_default_height is not None and data_row_indices:
        for r in data_row_indices: set_height(r, data_default_height, "data_default")
    before_footer_height = row_heights_cfg.get("before_footer")
    if before_footer_height is not None and row_before_footer_idx > 0: set_height(row_before_footer_idx, before_footer_height, "before_footer")

    # --- Footer Height Logic ---
    footer_height_config = row_heights_cfg.get("footer")
    match_header_height_flag = row_heights_cfg.get("footer_matches_header_height", True) # Default to True

    final_footer_height = None
    if match_header_height_flag and actual_header_height is not None:
        final_footer_height = actual_header_height # Use header height if flag is true and header height was set
    elif footer_height_config is not None:
        final_footer_height = footer_height_config # Otherwise, use specific footer height if defined

    if final_footer_height is not None and footer_row_index > 0:
        set_height(footer_row_index, final_footer_height, "footer")
    # --- End Footer Height Logic ---

    specific_heights = row_heights_cfg.get("specific_rows")
    if isinstance(specific_heights, dict):
        for row_str, height_val in specific_heights.items():
            try: row_num = int(row_str); set_height(row_num, height_val, f"specific_row_{row_num}")
            except ValueError: pass # Ignore invalid row numbers


# --- Main Execution Guard --- (Keep existing)
if __name__ == "__main__":
    print("invoice_utils.py executed directly.")
    pass